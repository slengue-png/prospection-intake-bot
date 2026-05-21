import os
import re
import json
import io
import base64
import zipfile
import hashlib
import datetime as dt
from typing import Dict, List, Any, Optional, Tuple

import requests
from PIL import Image
import pytesseract
import numpy as np
import cv2
import easyocr

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


# ============================================================
# ENV / CONFIG
# ============================================================

def env_str(name: str, default: str = "") -> str:
    v = os.environ.get(name, default)
    if v is None:
        return default
    return str(v).strip()


def env_int(name: str, default: int) -> int:
    s = env_str(name, "")
    if s == "":
        return default
    try:
        return int(s)
    except Exception:
        return default


def paris_ymd_fallback() -> str:
    return dt.date.today().strftime("%Y-%m-%d")


WORKER_BASE_URL = env_str("WORKER_BASE_URL")
EXPORT_TOKEN = env_str("EXPORT_TOKEN")
TELEGRAM_TOKEN = env_str("TELEGRAM_TOKEN")

BREVO_API_KEY = env_str("BREVO_API_KEY")
BREVO_SENDER_EMAIL = env_str("BREVO_SENDER_EMAIL", "no-reply@example.com")
BREVO_SENDER_NAME = env_str("BREVO_SENDER_NAME", "Prospection Bot")

MAIL_ROUTING_JSON = env_str("MAIL_ROUTING_JSON", "")

GOOGLE_PLACES_API_KEY = env_str("GOOGLE_PLACES_API_KEY", "")
GEMINI_API_KEY = env_str("GEMINI_API_KEY", "")

SEND_MODE = env_str("SEND_MODE", "individual").lower()
RUN_DATE = env_str("RUN_DATE", paris_ymd_fallback())
AGENCY = env_str("AGENCY", "").upper()
INITIALS = env_str("INITIALS", "").upper()

MAX_OCR_IMAGES = env_int("MAX_OCR_IMAGES", 50)
MAX_PHOTO_IMAGES = env_int("MAX_PHOTO_IMAGES", 15)

OUT_DIR = env_str("OUT_DIR", "out").strip() or "out"
if os.path.exists(OUT_DIR) and not os.path.isdir(OUT_DIR):
    print(f"[WARN] OUT_DIR='{OUT_DIR}' existe mais n'est pas un dossier. Fallback -> 'exports'")
    OUT_DIR = "exports"
os.makedirs(OUT_DIR, exist_ok=True)

VALID_AGENCIES = {"GR", "VR", "GRS", "SLS"}

EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")
EMAIL_IN_TEXT_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
PHONE_RE = re.compile(r"(?:(?:\+33|0)[\s\.-]*[1-9](?:[\s\.-]*\d{2}){4})")
URL_RE = re.compile(r"(https?://[^\s)]+|www\.[^\s)]+)", re.I)
CP_RE = re.compile(r"\b((?:0[1-9]|[1-8]\d|9[0-5])\s?\d{3}|97\d{3}|98\d{3})\b")

DISPOSABLE_DOMAINS = {
    "tempmail.com",
    "guerrillamail.com",
    "yopmail.com",
    "10minutemail.com",
}

FREE_EMAIL_DOMAINS = {
    "gmail.com", "outlook.com", "hotmail.com", "yahoo.com", "yahoo.fr",
    "icloud.com", "free.fr", "orange.fr", "laposte.net", "live.fr"
}

HEADERS = [
    "date", "agency", "initials", "name", "address", "postal_code", "city",
    "siret", "naf", "activity_summary", "dirigeant", "interlocuteur",
    "contact_firstname", "contact_lastname", "phone", "phone2", "email",
    "website", "resume", "commande"
]


# ============================================================
# ROUTING
# ============================================================

DEFAULT_ROUTING = {
    "agencies": {
        "GR": {
            "manager": {"initials": "JL", "email": "jennifer.laurens@ras-interim.fr"},
            "commercial": {"initials": "CZ", "email": "celine.zunarelli@ras-interim.fr"},
        },
        "VR": {
            "manager": {"initials": "JB", "email": "jelena.carrasso@ras-interim.fr"},
            "commercial": {"initials": "LB", "email": "laura.berthet@ras-interim.fr"},
        },
        "GRS": {
            "manager": {"initials": "PV", "email": "pauline.vieira@ras-interim.fr"},
            "commercial": {"initials": "ST", "email": "severine.thevenin@ras-interim.fr"},
        },
        "SLS": {
            "manager": {"initials": "AC", "email": "aurelie.curt@ras-interim.fr"},
            "commercial": {"initials": "AC", "email": "aurelie.curt@ras-interim.fr"},
        },
    },
    "users": {
        "JL": "jennifer.laurens@ras-interim.fr",
        "CZ": "celine.zunarelli@ras-interim.fr",
        "JB": "jelena.carrasso@ras-interim.fr",
        "LB": "laura.berthet@ras-interim.fr",
        "PV": "pauline.vieira@ras-interim.fr",
        "ST": "severine.thevenin@ras-interim.fr",
        "AC": "aurelie.curt@ras-interim.fr",
        "SL": "samuel.lengue@ras-interim.fr",
    },
    "admin": {"initials": "SL", "email": "samuel.lengue@ras-interim.fr"},
}


def load_routing() -> Dict[str, Any]:
    if MAIL_ROUTING_JSON:
        try:
            data = json.loads(MAIL_ROUTING_JSON)
            if isinstance(data, dict) and data:
                return data
        except Exception as e:
            print(f"[WARN] MAIL_ROUTING_JSON invalid JSON, fallback default. err={e}")
    return DEFAULT_ROUTING


ROUTING = load_routing()


def clean_email(v) -> str:
    if v is None:
        return ""
    if isinstance(v, dict):
        v = v.get("email") or ""
    if not isinstance(v, str):
        v = str(v)
    v = v.strip().lower()
    v = re.sub(r"\s+", "", v)
    return v


def is_valid_email(s: str) -> bool:
    s = clean_email(s)
    if not s or len(s) > 254:
        return False
    if not EMAIL_RE.match(s):
        return False
    domain = s.split("@", 1)[1]
    if domain in DISPOSABLE_DOMAINS:
        return False
    return True


def email_for_initials(initials: str) -> Optional[str]:
    initials = (initials or "").upper().strip()
    if not initials:
        return None

    users = ROUTING.get("users") or {}
    if initials in users:
        em = clean_email(users.get(initials))
        if is_valid_email(em):
            return em

    agencies = ROUTING.get("agencies") or {}
    for _, cfg in agencies.items():
        for role in ("manager", "commercial"):
            r = (cfg or {}).get(role) or {}
            if (r.get("initials") or "").upper() == initials:
                em2 = clean_email(r.get("email") or "")
                return em2 if is_valid_email(em2) else None
    return None


# ============================================================
# RUNTIME STATS / CACHE
# ============================================================

STATS = {
    "worker_dump_calls": 0,
    "telegram_getfile_calls": 0,
    "telegram_file_downloads": 0,
    "gemini_vision_calls": 0,
    "gemini_activity_calls": 0,
    "gouv_calls": 0,
    "places_calls": 0,
    "scrape_calls": 0,
    "emails_sent": 0,
    "ocr_easyocr_calls": 0,
    "ocr_tesseract_calls": 0,
    "dedupe_groups": 0,
    "dedupe_rows_merged": 0,
    "linked_cards_processed": 0,
    "linked_cards_missing": 0,
    "lot_cards_promoted": 0,
    "lot_photos_promoted": 0,
    "fusion_logs": 0,
}

GOUV_CACHE: Dict[str, Dict[str, Any]] = {}
PLACES_CACHE: Dict[str, Dict[str, str]] = {}
EMAIL_SCRAPE_CACHE: Dict[str, str] = {}
GEMINI_ACTIVITY_CACHE: Dict[str, str] = {}
GEMINI_CARD_CACHE: Dict[str, Dict[str, str]] = {}
GEMINI_FACADE_CACHE: Dict[str, Dict[str, str]] = {}
TG_FILE_PATH_CACHE: Dict[str, str] = {}
TG_FILE_BYTES_CACHE: Dict[str, bytes] = {}
CARD_PIPELINE_CACHE: Dict[str, Dict[str, Any]] = {}
PHOTO_PIPELINE_CACHE: Dict[str, Dict[str, Any]] = {}
OCR_TEXT_CACHE: Dict[str, str] = {}


# ============================================================
# EASYOCR READER
# ============================================================

EASYOCR_READER = None


def get_easyocr_reader():
    global EASYOCR_READER
    if EASYOCR_READER is None:
        try:
            EASYOCR_READER = easyocr.Reader(["fr", "en"], gpu=False)
        except Exception as e:
            print(f"[WARN] EasyOCR unavailable: {e}")
            EASYOCR_READER = False
    return EASYOCR_READER


# ============================================================
# HELPERS GÉNÉRAUX
# ============================================================

def safe_strip(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip())


def normalize_spaces(v: str) -> str:
    return re.sub(r"\s+", " ", (v or "").strip())


def normalize_upper(v: str) -> str:
    return normalize_spaces(v).upper()


def normalize_cp(cp: str) -> str:
    return re.sub(r"\s+", "", (cp or "").strip())


def sha1_text(s: str) -> str:
    return hashlib.sha1((s or "").encode("utf-8", errors="ignore")).hexdigest()


def image_hash_bytes(img_bytes: bytes) -> str:
    return hashlib.sha1(img_bytes or b"").hexdigest()


def domain_from_email(email: str) -> str:
    email = (email or "").strip().lower()
    if "@" not in email:
        return ""
    dom = email.split("@", 1)[1].strip()
    dom = re.sub(r"[^a-z0-9.\-]", "", dom)
    if dom in FREE_EMAIL_DOMAINS:
        return ""
    return dom


def domain_from_url(url: str) -> str:
    if not url:
        return ""
    u = url.lower().strip()
    u = re.sub(r"^https?://", "", u)
    u = u.split("/", 1)[0]
    u = u.split(":", 1)[0]
    u = re.sub(r"[^a-z0-9.\-]", "", u)
    if u.startswith("www."):
        u = u[4:]
    return u


def brand_from_domain(dom: str) -> str:
    dom = (dom or "").strip().lower()
    if not dom:
        return ""
    dom = dom.split(".", 1)[0]
    dom = dom.replace("-", " ").replace("_", " ")
    dom = re.sub(r"\s+", " ", dom).strip()
    return dom


def normalize_text(s: str) -> str:
    s = (s or "").lower().strip()
    s = s.replace("é", "e").replace("è", "e").replace("ê", "e").replace("ë", "e")
    s = s.replace("à", "a").replace("â", "a")
    s = s.replace("î", "i").replace("ï", "i")
    s = s.replace("ô", "o")
    s = s.replace("ù", "u").replace("û", "u").replace("ü", "u")
    s = s.replace("ç", "c")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_company_name(s: str) -> str:
    s = normalize_text(s)
    stop = {
        "sas", "sasu", "sa", "sarl", "eurl", "scop", "scp", "snc", "sci",
        "societe", "compagnie", "groupe", "generale", "france", "holding", "services"
    }
    parts = [p for p in s.split() if p not in stop]
    return " ".join(parts).strip()


def validate_siret(raw: str) -> str:
    siret = re.sub(r"\D", "", (raw or "").strip())
    if len(siret) != 14 or not siret.isdigit():
        if siret:
            print(f"[WARN] SIRET invalid length/non-digit: {raw}")
        return ""

    total = 0
    for i in range(14):
        digit = int(siret[13 - i])
        if i % 2 == 1:
            digit *= 2
        if digit > 9:
            digit -= 9
        total += digit

    if total % 10 == 0:
        return siret

    print(f"[WARN] SIRET invalid Luhn: {raw}")
    return ""


def normalize_phone(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    s = re.sub(r"[^\d+]", "", s)
    if s.startswith("+33"):
        s = "0" + s[3:]
    if not re.fullmatch(r"0\d{9}", s or ""):
        return ""
    return s


def extract_all_phones(text: str) -> List[str]:
    if not text:
        return []
    out = []
    seen = set()
    for m in PHONE_RE.finditer(text):
        p = normalize_phone(m.group(0))
        if not p:
            continue
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def split_phones_by_priority(phones: List[str]) -> Tuple[str, str]:
    fixed = []
    mobile = []
    other = []

    for p in phones:
        pn = normalize_phone(p)
        if not pn:
            continue
        if pn.startswith(("01", "02", "03", "04", "05")):
            if pn not in fixed:
                fixed.append(pn)
        elif pn.startswith(("06", "07")):
            if pn not in mobile:
                mobile.append(pn)
        else:
            if pn not in other:
                other.append(pn)

    phone = ""
    phone2 = ""

    if fixed:
        phone = fixed[0]
        if mobile:
            phone2 = mobile[0]
        elif len(fixed) > 1:
            phone2 = fixed[1]
        elif other:
            phone2 = other[0]
    elif mobile:
        phone = mobile[0]
        if len(mobile) > 1:
            phone2 = mobile[1]
        elif other:
            phone2 = other[0]
    elif other:
        phone = other[0]
        if len(other) > 1:
            phone2 = other[1]

    if phone and phone2 and phone == phone2:
        phone2 = ""

    return phone, phone2


def best_email(text: str) -> str:
    if not text:
        return ""
    candidates = [clean_email(m.group(0)) for m in EMAIL_IN_TEXT_RE.finditer(text)]
    for em in candidates:
        if is_valid_email(em):
            return em
    return ""


def extract_all_emails(text: str) -> List[str]:
    if not text:
        return []
    out = []
    seen = set()
    for m in EMAIL_IN_TEXT_RE.finditer(text):
        em = clean_email(m.group(0))
        if is_valid_email(em) and em not in seen:
            seen.add(em)
            out.append(em)
    return out


def extract_urls(text: str) -> List[str]:
    if not text:
        return []
    out, seen = [], set()
    for m in URL_RE.finditer(text):
        u = m.group(0).strip()
        if u.lower().startswith("www."):
            u = "https://" + u
        u = u.rstrip(".,;:")
        k = u.lower()
        if k not in seen:
            seen.add(k)
            out.append(u)
    return out


def extract_postal_code(text: str) -> str:
    if not text:
        return ""
    cps = CP_RE.findall(text)
    return normalize_cp(cps[0]) if cps else ""


def split_human_name(full: str) -> Tuple[str, str]:
    s = normalize_spaces(full)
    if not s:
        return "", ""
    parts = s.split(" ")
    if len(parts) == 1:
        return "", parts[0]
    return parts[0], " ".join(parts[1:])


def ensure_contact_fallback(d: Dict[str, Any]) -> Dict[str, Any]:
    fn = (d.get("contact_firstname") or "").strip()
    ln = (d.get("contact_lastname") or "").strip()
    interlocuteur = (d.get("interlocuteur") or "").strip()
    dirigeant = (d.get("dirigeant") or "").strip()

    if interlocuteur:
        f, l = split_human_name(interlocuteur)
        if not fn and f:
            d["contact_firstname"] = f
        if not ln and l:
            d["contact_lastname"] = l

    fn = (d.get("contact_firstname") or "").strip()
    ln = (d.get("contact_lastname") or "").strip()

    if (not fn and not ln) and dirigeant:
        f, l = split_human_name(dirigeant)
        d["contact_firstname"] = f or ""
        d["contact_lastname"] = l or (dirigeant or "")

    if not (d.get("interlocuteur") or "").strip():
        combo = f"{(d.get('contact_firstname') or '').strip()} {(d.get('contact_lastname') or '').strip()}".strip()
        if combo:
            d["interlocuteur"] = combo
        elif dirigeant:
            d["interlocuteur"] = dirigeant

    return d


def strip_postal_city_from_address(addr: str, postal_code: str, city: str) -> str:
    addr = (addr or "").strip()
    if not addr:
        return ""

    pc = (postal_code or "").strip()
    ct = (city or "").strip()

    addr = re.sub(r"\s+", " ", addr)

    if pc and ct:
        tail = f"{pc} {ct}".strip()
        addr = re.sub(rf"(?:\s|-)?{re.escape(tail)}\s*$", "", addr, flags=re.IGNORECASE).strip()

    if pc:
        addr = re.sub(rf"(?:\s|-)?{re.escape(pc)}\s*$", "", addr).strip()

    addr = re.sub(r"\s+", " ", addr).strip()
    return addr


def normalize_row_address(row: Dict[str, Any]) -> Dict[str, Any]:
    row["address"] = strip_postal_city_from_address(
        row.get("address", ""),
        row.get("postal_code", ""),
        row.get("city", ""),
    )
    return row


def uppercase_business_fields(row: Dict[str, Any]) -> Dict[str, Any]:
    for k in ["name", "address", "city", "dirigeant", "interlocuteur"]:
        row[k] = normalize_upper(row.get(k, ""))
    return row


def log_fusion(reason: str, payload: Dict[str, Any]) -> None:
    STATS["fusion_logs"] += 1
    try:
        print(json.dumps({"fusion_reason": reason, **payload}, ensure_ascii=False))
    except Exception:
        print(f"[FUSION] {reason} | {payload}")


def autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > 80:
                s = s[:80]
            max_len = max(max_len, len(s))
            max_len = min(max_len, 60)
        ws.column_dimensions[get_column_letter(col)].width = max_len


# ============================================================
# HTTP HELPERS
# ============================================================

def worker_dump(kind: str, date: str) -> List[Dict[str, Any]]:
    if not WORKER_BASE_URL or not EXPORT_TOKEN:
        raise RuntimeError("Missing WORKER_BASE_URL / EXPORT_TOKEN")
    STATS["worker_dump_calls"] += 1
    url = f"{WORKER_BASE_URL.rstrip('/')}/dump?date={date}&kind={kind}"
    r = requests.get(url, headers={"X-Export-Token": EXPORT_TOKEN}, timeout=45)
    if r.status_code != 200:
        raise RuntimeError(f"/dump failed {kind} {r.status_code}: {r.text[:400]}")
    lines = [ln.strip() for ln in r.text.splitlines() if ln.strip()]
    out: List[Dict[str, Any]] = []
    for ln in lines:
        try:
            out.append(json.loads(ln))
        except Exception:
            pass
    return out


def tg_get_file_path(file_id: str) -> Optional[str]:
    if not file_id:
        return None
    if file_id in TG_FILE_PATH_CACHE:
        return TG_FILE_PATH_CACHE[file_id]
    if not TELEGRAM_TOKEN:
        raise RuntimeError("Missing TELEGRAM_TOKEN")

    STATS["telegram_getfile_calls"] += 1
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getFile"
    r = requests.post(url, json={"file_id": file_id}, timeout=25)
    r.raise_for_status()
    j = r.json()
    if not j.get("ok"):
        return None
    fp = j["result"]["file_path"]
    TG_FILE_PATH_CACHE[file_id] = fp
    return fp


def tg_download_file(file_id: str) -> bytes:
    if not file_id:
        return b""
    if file_id in TG_FILE_BYTES_CACHE:
        return TG_FILE_BYTES_CACHE[file_id]
    file_path = tg_get_file_path(file_id)
    if not file_path:
        return b""
    STATS["telegram_file_downloads"] += 1
    r = requests.get(f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}", timeout=60)
    r.raise_for_status()
    content = r.content
    TG_FILE_BYTES_CACHE[file_id] = content
    return content


# ============================================================
# OCR HYBRIDE
# ============================================================

def pil_to_cv(img: Image.Image) -> np.ndarray:
    arr = np.array(img)
    if len(arr.shape) == 2:
        return arr
    return cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)


def preprocess_image_variants(img_bytes: bytes) -> List[Image.Image]:
    variants: List[Image.Image] = []
    try:
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    except Exception:
        return variants

    variants.append(img)

    try:
        big = img.resize((img.width * 2, img.height * 2))
        variants.append(big)
    except Exception:
        pass

    try:
        gray = img.convert("L")
        arr = np.array(gray)
        arr = cv2.equalizeHist(arr)
        variants.append(Image.fromarray(arr))
    except Exception:
        pass

    try:
        gray = img.convert("L")
        arr = np.array(gray)
        _, th = cv2.threshold(arr, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        variants.append(Image.fromarray(th))
    except Exception:
        pass

    try:
        gray = np.array(img.convert("L"))
        ad = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 15)
        variants.append(Image.fromarray(ad))
    except Exception:
        pass

    return variants


def crop_image_zones(img: Image.Image) -> List[Image.Image]:
    w, h = img.size
    zones = [img]
    zones.append(img.crop((0, 0, w, h // 3)))
    zones.append(img.crop((0, h // 3, w, 2 * h // 3)))
    zones.append(img.crop((0, 2 * h // 3, w, h)))
    zones.append(img.crop((0, 0, w // 2, h)))
    zones.append(img.crop((w // 2, 0, w, h)))
    return zones


def ocr_easyocr_on_pil(img: Image.Image) -> str:
    try:
        reader = get_easyocr_reader()
        if not reader:
            return ""
        STATS["ocr_easyocr_calls"] += 1
        arr = pil_to_cv(img)
        results = reader.readtext(arr, detail=0, paragraph=False)
        return "\n".join([str(x).strip() for x in results if str(x).strip()])
    except Exception:
        return ""


def ocr_tesseract_on_pil(img: Image.Image) -> str:
    try:
        STATS["ocr_tesseract_calls"] += 1
        txt1 = pytesseract.image_to_string(img, lang="fra+eng", config="--psm 6") or ""
        txt2 = pytesseract.image_to_string(img, lang="fra+eng", config="--psm 11") or ""
        return "\n".join([t for t in [txt1, txt2] if safe_strip(t)])
    except Exception:
        return ""


def merge_text_blocks(texts: List[str]) -> str:
    lines = []
    seen = set()
    for txt in texts:
        for ln in (txt or "").splitlines():
            ln = re.sub(r"\s+", " ", ln).strip()
            if not ln:
                continue
            key = ln.lower()
            if key not in seen:
                seen.add(key)
                lines.append(ln)
    return "\n".join(lines)


def ocr_image_bytes(img_bytes: bytes, light: bool = False) -> str:
    if not img_bytes:
        return ""

    cache_key = f"ocr:{image_hash_bytes(img_bytes)}:{'light' if light else 'full'}"
    if cache_key in OCR_TEXT_CACHE:
        return OCR_TEXT_CACHE[cache_key]

    variants = preprocess_image_variants(img_bytes)
    all_texts = []

    base_variants = variants[: (2 if light else len(variants))]

    for base_img in base_variants:
        zones = crop_image_zones(base_img)
        for zone in zones[: (2 if light else len(zones))]:
            # EasyOCR + Tesseract, les deux restent présents
            t1 = ocr_easyocr_on_pil(zone)
            if t1:
                all_texts.append(t1)
            t2 = ocr_tesseract_on_pil(zone)
            if t2:
                all_texts.append(t2)

    merged = merge_text_blocks(all_texts)
    OCR_TEXT_CACHE[cache_key] = merged
    return merged


# ============================================================
# GEMINI
# ============================================================

def _guess_mime(img_bytes: bytes) -> str:
    try:
        im = Image.open(io.BytesIO(img_bytes))
        fmt = (im.format or "").upper()
        if fmt == "PNG":
            return "image/png"
        if fmt in ("JPG", "JPEG"):
            return "image/jpeg"
    except Exception:
        pass
    return "image/jpeg"


def gemini_vision_json(img_bytes: bytes, prompt: str, max_tokens: int = 650) -> Dict[str, Any]:
    if not GEMINI_API_KEY or not img_bytes:
        return {}

    endpoint = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent"
    mime = _guess_mime(img_bytes)
    b64 = base64.b64encode(img_bytes).decode("ascii")

    payload = {
        "contents": [{
            "parts": [
                {"text": prompt},
                {"inlineData": {"mimeType": mime, "data": b64}}
            ]
        }],
        "generationConfig": {"temperature": 0.1, "maxOutputTokens": max_tokens}
    }

    try:
        STATS["gemini_vision_calls"] += 1
        r = requests.post(endpoint, params={"key": GEMINI_API_KEY}, json=payload, timeout=60)
        if r.status_code >= 300:
            return {}
        j = r.json()
        parts = (((j.get("candidates") or [None])[0] or {}).get("content") or {}).get("parts") or []
        raw = ""
        for p in parts:
            if isinstance(p, dict) and p.get("text"):
                raw += p["text"]
        raw = (raw or "").strip()
        raw = re.sub(r"^```json", "", raw, flags=re.I).strip()
        raw = re.sub(r"```$", "", raw).strip()
        m = re.search(r"\{.*\}", raw, re.S)
        if not m:
            return {}
        return json.loads(m.group(0))
    except Exception:
        return {}


def gemini_extract_business_card(img_bytes: bytes) -> Dict[str, str]:
    if not img_bytes:
        return {}
    key = image_hash_bytes(img_bytes)
    if key in GEMINI_CARD_CACHE:
        return GEMINI_CARD_CACHE[key]

    prompt = (
        "Tu es un extracteur de carte de visite. "
        "Lis l'IMAGE et retourne STRICTEMENT un JSON avec les clés : "
        "name, company, title, email, phone, website, postal_code, city, address. "
        "Règles : si inconnu mettre chaîne vide ; email en minuscules ; "
        "postal_code = 5 chiffres si visible ; "
        "name = prénom + nom si visibles ; "
        "address = adresse la plus propre possible."
    )
    d = gemini_vision_json(img_bytes, prompt, max_tokens=700) or {}
    out = {
        "name": safe_strip(d.get("name", "")),
        "company": safe_strip(d.get("company", "")),
        "title": safe_strip(d.get("title", "")),
        "email": clean_email(d.get("email", "")),
        "phone": normalize_phone(d.get("phone", "")),
        "website": safe_strip(d.get("website", "")),
        "postal_code": normalize_cp(d.get("postal_code", "")),
        "city": safe_strip(d.get("city", "")),
        "address": safe_strip(d.get("address", "")),
    }
    GEMINI_CARD_CACHE[key] = out
    return out


def gemini_extract_facade_logo(img_bytes: bytes) -> Dict[str, str]:
    if not img_bytes:
        return {}
    key = image_hash_bytes(img_bytes)
    if key in GEMINI_FACADE_CACHE:
        return GEMINI_FACADE_CACHE[key]

    prompt = (
        "Tu analyses une photo de prospection (façade, enseigne, logo). "
        "Retourne STRICTEMENT un JSON avec les clés : company, city. "
        "Règles : si inconnu mettre chaîne vide."
    )
    d = gemini_vision_json(img_bytes, prompt, max_tokens=350) or {}
    out = {
        "company": safe_strip(d.get("company", "")),
        "city": safe_strip(d.get("city", "")),
    }
    GEMINI_FACADE_CACHE[key] = out
    return out


def gemini_activity_summary(name: str, naf: str, website: str, company_hint: str = "") -> str:
    if not GEMINI_API_KEY:
        return ""

    cache_key = f"{name}|{naf}|{website}|{company_hint}"
    if cache_key in GEMINI_ACTIVITY_CACHE:
        return GEMINI_ACTIVITY_CACHE[cache_key]

    if not (name or naf or website):
        return ""

    endpoint = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent"
    prompt = (
        "Tu es un normaliseur d'activité d'entreprise. "
        "Retourne STRICTEMENT un JSON avec la clé EXACTE: activity_summary. "
        "Règles : activité très courte, factuelle, maximum 60 caractères, pas de marketing, si inconnu vide. "
        f"Nom entreprise: {name}\n"
        f"Code NAF: {naf}\n"
        f"Site web: {website}\n"
        f"Indice complémentaire: {company_hint}\n"
    )

    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.1, "maxOutputTokens": 120}
    }

    try:
        STATS["gemini_activity_calls"] += 1
        r = requests.post(endpoint, params={"key": GEMINI_API_KEY}, json=payload, timeout=30)
        if r.status_code >= 300:
            return ""
        j = r.json()
        parts = (((j.get("candidates") or [None])[0] or {}).get("content") or {}).get("parts") or []
        raw = ""
        for p in parts:
            if isinstance(p, dict) and p.get("text"):
                raw += p["text"]
        raw = (raw or "").strip()
        raw = re.sub(r"^```json", "", raw, flags=re.I).strip()
        raw = re.sub(r"```$", "", raw).strip()
        m = re.search(r"\{.*\}", raw, re.S)
        if not m:
            return ""
        d = json.loads(m.group(0))
        out = safe_strip(d.get("activity_summary", ""))[:60]
        GEMINI_ACTIVITY_CACHE[cache_key] = out
        return out
    except Exception:
        return ""


# ============================================================
# HEURISTIQUES OCR / ENTREPRISE / PERSONNE / ADRESSE
# ============================================================

def is_probable_person_name(s: str) -> bool:
    s = re.sub(r"\s+", " ", (s or "").strip())
    if not s:
        return False
    low = normalize_text(s)
    if "@" in s or "www" in low:
        return False
    if re.search(r"\d", s):
        return False

    bad_exact = {"alt gr", "ait gr", "alt qr", "ait qr", "cs", "sas", "sarl", "sa"}
    if low in bad_exact:
        return False

    bad_tokens = {
        "restaurant", "responsable", "commerce", "collectivite", "territoire",
        "recyclage", "valorisation", "dechets", "drh", "developpement", "rh",
        "onyx", "auvergne", "rhone", "alpes", "sas", "sarl", "eurl", "sa",
        "service", "environnement", "gestion", "cedex", "centr", "alp", "cs",
        "experience", "luxe", "scannez", "scanner", "savoir", "plus", "aqualone"
    }
    words = low.split()
    if len(words) < 2 or len(words) > 4:
        return False
    if any(w in bad_tokens for w in words):
        return False

    return sum(1 for w in words if len(w) >= 2) >= 2


def clean_ocr_line(line: str) -> str:
    s = re.sub(r"\s+", " ", (line or "").strip())
    s = s.replace("|", " ").replace("•", " ").replace("—", " ").replace("–", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def clean_ocr_lines(text: str) -> List[str]:
    bad_contains = [
        "scannez-moi",
        "scannez moi",
        "en savoir plus",
        "qr code",
    ]
    out = []
    seen = set()

    for raw in (text or "").splitlines():
        ln = clean_ocr_line(raw)
        if not ln:
            continue

        low = normalize_text(ln)
        if any(x in low for x in bad_contains):
            continue

        key = low
        if key in seen:
            continue
        seen.add(key)
        out.append(ln)

    return out


def looks_like_company_line(line: str) -> bool:
    ln = clean_ocr_line(line)
    if not ln:
        return False

    low = normalize_text(ln)

    if "@" in ln:
        return False
    if URL_RE.search(ln):
        return False
    if PHONE_RE.search(ln):
        return False
    if CP_RE.search(ln):
        return False

    bad_words = {
        "directeur", "responsable", "rh", "site", "manager",
        "portable", "port", "tel", "telephone", "mail", "email",
        "scannez", "scanner", "savoir", "experience", "luxe"
    }
    words = set(low.split())
    if words & bad_words:
        return False

    if len(ln) < 3:
        return False
    if len(ln) > 50:
        return False

    return True


def detect_company_from_lines(lines: List[str], gemini_company: str = "", website: str = "", email: str = "") -> str:
    if gemini_company and len(gemini_company.strip()) >= 3:
        return re.sub(r"\s+", " ", gemini_company).strip()

    dom_web = brand_from_domain(domain_from_url(website))
    dom_email = brand_from_domain(domain_from_email(email))

    candidates = []
    for i, ln in enumerate(lines[:8]):
        if not looks_like_company_line(ln):
            continue

        low = normalize_company_name(ln)
        score = 0

        if i <= 1:
            score += 25
        if ln == ln.upper():
            score += 20
        if len(ln.split()) <= 3:
            score += 15
        if len(ln) <= 20:
            score += 10
        if dom_web and normalize_company_name(dom_web) in low:
            score += 30
        if dom_email and normalize_company_name(dom_email) in low:
            score += 30
        if len(ln.split()) in [2, 3]:
            score += 10

        candidates.append((score, ln))

    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1].strip()

    for cand in [dom_web, dom_email]:
        cand = re.sub(r"\s+", " ", (cand or "").strip())
        if len(cand) >= 3:
            return cand

    return ""


def detect_person_name_from_lines(lines: List[str], company_line: str = "") -> str:
    company_n = normalize_company_name(company_line)

    bad_person_phrases = {
        "le bois essence",
        "de l experience du luxe",
        "scannez moi pour en savoir plus",
        "centre aquatique aqualone",
    }

    for ln in lines[:12]:
        low = normalize_text(ln)

        if low in bad_person_phrases:
            continue

        if any(x in low for x in ["experience", "luxe", "scannez", "savoir plus"]):
            continue

        if not is_probable_person_name(ln):
            continue

        if company_n and normalize_company_name(ln) == company_n:
            continue

        return ln.strip()

    return ""


def clean_address_candidate(addr: str) -> str:
    a = re.sub(r"\s+", " ", (addr or "").strip())
    if not a:
        return ""

    low = normalize_text(a)

    if "@" in a:
        return ""
    if URL_RE.search(a):
        return ""
    if PHONE_RE.search(a):
        return ""
    if low.startswith("port"):
        return ""
    if low.startswith("tel"):
        return ""
    if "portable" in low:
        return ""

    return a


def extract_local_address_from_text(text: str) -> Tuple[str, str, str]:
    if not text:
        return "", "", ""

    lines = [re.sub(r"\s+", " ", ln).strip() for ln in text.splitlines()]
    lines = [ln for ln in lines if ln]

    best_addr = ""
    best_cp = ""
    best_city = ""

    for i, ln in enumerate(lines):
        m = CP_RE.search(ln)
        if not m:
            continue

        cp = normalize_cp(m.group(1))
        before = ln[:m.start()].strip(" -,:;")
        after = ln[m.end():].strip(" -,:;")

        addr_line = ln
        if len(before) < 4 and i > 0:
            prev = lines[i - 1]
            if not CP_RE.search(prev):
                addr_line = f"{prev} {ln}".strip()

        city = after[:60].strip()

        addr_line = re.sub(r"\s+", " ", addr_line).strip()
        city = re.sub(r"\s+", " ", city).strip()

        if cp:
            best_addr, best_cp, best_city = addr_line, cp, city
            break

    return best_addr, best_cp, best_city


def extract_best_company_hint(lines: List[str], gemini_company: str, website: str, email: str) -> str:
    company = detect_company_from_lines(lines, gemini_company=gemini_company, website=website, email=email)
    return deduce_company(company, email, website)


def deduce_company(company_raw: str, email: str, website: str) -> str:
    company_raw = (company_raw or "").strip()
    dom_email = brand_from_domain(domain_from_email(email))
    dom_web = brand_from_domain(domain_from_url(website))

    for cand in [company_raw, dom_web, dom_email]:
        cand = re.sub(r"\s+", " ", (cand or "").strip())
        if len(cand) >= 4:
            return cand

    return company_raw


def choose_final_company_name(
    gouv_name: str,
    gemini_company: str,
    detected_company_line: str,
    person_name: str,
    website: str,
    email: str
) -> str:
    person_n = normalize_company_name(person_name)
    dom_web = brand_from_domain(domain_from_url(website))
    dom_mail = brand_from_domain(domain_from_email(email))

    candidates = [
        re.sub(r"\s+", " ", (gouv_name or "").strip()),
        re.sub(r"\s+", " ", (gemini_company or "").strip()),
        re.sub(r"\s+", " ", (detected_company_line or "").strip()),
    ]

    scored = []

    for cand in candidates:
        if not cand:
            continue

        cand_norm = normalize_company_name(cand)
        if not cand_norm:
            continue

        if person_n and cand_norm == person_n:
            continue
        if is_probable_person_name(cand):
            continue

        score = 0

        if gouv_name and cand.strip() == gouv_name.strip():
            score += 100

        if dom_web and normalize_company_name(dom_web) in cand_norm:
            score += 30
        if dom_mail and normalize_company_name(dom_mail) in cand_norm:
            score += 30

        wc = len(cand.split())
        if wc >= 2:
            score += 15
        if wc == 1:
            score -= 10

        if len(cand) <= 5:
            score -= 15

        scored.append((score, cand))

    if scored:
        scored.sort(key=lambda x: x[0], reverse=True)
        return scored[0][1].strip()

    for cand in [dom_web, dom_mail]:
        cand = re.sub(r"\s+", " ", (cand or "").strip())
        if len(cand) >= 3:
            return cand

    return ""


def infer_activity_fallback(name: str, website: str = "", naf: str = "") -> str:
    n = normalize_text(name)
    w = normalize_text(domain_from_url(website))
    z = f"{n} {w} {naf}".strip()

    rules = [
        (["atelier delaye", "menuiserie delaye", "menuiserie", "bois"], "MENUISERIE"),
        (["vert marine", "aqualone", "aquatique", "piscine"], "CENTRE AQUATIQUE"),
        (["groupe cheval", "cheval molina", "transport", "travaux publics", "benne"], "TRANSPORT / TRAVAUX PUBLICS"),
        (["serrurerie boret", "boret", "serrurerie"], "SERRURERIE / MÉTALLERIE"),
        (["ad resine", "resine", "revetement"], "RÉSINE / REVÊTEMENTS"),
        (["euromaster", "pneu", "garage"], "PNEUMATIQUES / ENTRETIEN AUTO"),
        (["lely", "environnement", "dechets"], "ENVIRONNEMENT / DECHETS"),
        (["keolis", "transport voyageurs"], "TRANSPORT DE VOYAGEURS"),
    ]

    for keys, label in rules:
        if any(k in z for k in keys):
            return label

    naf_map = {
        "4332B": "MENUISERIE",
        "5610C": "RESTAURATION",
        "3811Z": "COLLECTE DE DÉCHETS",
        "4531Z": "PNEUMATIQUES / AUTO",
        "4333Z": "REVÊTEMENTS / RÉSINE",
        "4399D": "TRAVAUX SPÉCIALISÉS",
        "2512Z": "SERRURERIE / MÉTALLERIE",
    }
    if naf in naf_map:
        return naf_map[naf]

    return ""


def enrich_activity_summary(name: str, naf: str, website: str, company_hint: str = "") -> str:
    val = gemini_activity_summary(name=name, naf=naf, website=website, company_hint=company_hint)
    if val:
        return val
    return infer_activity_fallback(name=name, website=website, naf=naf)


# ============================================================
# ENRICHISSEMENT ENTREPRISE
# ============================================================

def search_gouv_company(name: str, city: str = "") -> Dict[str, str]:
    name = (name or "").strip()
    if not name:
        return {}

    cache_key = f"search_gouv_company|{name}|{city}"
    if cache_key in GOUV_CACHE:
        return dict(GOUV_CACHE[cache_key])

    try:
        STATS["gouv_calls"] += 1
        q = name if not city else f"{name} {city}".strip()
        url = f"https://recherche-entreprises.api.gouv.fr/search?q={requests.utils.quote(q)}&page=1&per_page=1"
        r = requests.get(url, headers={"accept": "application/json"}, timeout=20)
        r.raise_for_status()
        j = r.json()
        res = (j.get("results") or [])
        if not res:
            GOUV_CACHE[cache_key] = {}
            return {}
        e = res[0]
        siege = e.get("siege") or {}

        dirigeant = ""
        arr = e.get("dirigeants") or e.get("representants") or []
        if arr:
            first = arr[0]
            if isinstance(first, str):
                dirigeant = first
            elif isinstance(first, dict):
                p = first.get("personne") or first
                prenom = p.get("prenom") or p.get("prenoms") or ""
                nom = p.get("nom") or p.get("nom_usage") or ""
                dirigeant = f"{prenom} {nom}".strip()

        naf = (e.get("activite_principale") or e.get("naf") or "").replace(".", "").replace(" ", "").upper()
        out = {
            "name": e.get("nom_raison_sociale") or e.get("denomination") or name,
            "siret": validate_siret(siege.get("siret") or ""),
            "naf": naf,
            "address": (siege.get("adresse") or siege.get("libelle_voie") or ""),
            "postal_code": normalize_cp(siege.get("code_postal") or ""),
            "city": (siege.get("libelle_commune") or city),
            "dirigeant": dirigeant,
        }
        GOUV_CACHE[cache_key] = out
        return dict(out)
    except Exception:
        return {}


def build_company_candidates(company_raw: str, email: str, website: str, city: str, ocr_text: str) -> List[str]:
    cands = []
    raw = (company_raw or "").strip()
    dom_email = brand_from_domain(domain_from_email(email))
    dom_web = brand_from_domain(domain_from_url(website))

    for x in [raw, dom_email, dom_web]:
        x = re.sub(r"\s+", " ", (x or "").strip())
        if x and x not in cands:
            cands.append(x)

    base = cands[:]
    for x in base:
        if city:
            y = f"{x} {city}".strip()
            if y not in cands:
                cands.append(y)

    if ocr_text:
        lines = [re.sub(r"\s+", " ", ln).strip() for ln in ocr_text.splitlines()]
        normalized_existing = {normalize_company_name(c) for c in cands}
        for ln in lines[:8]:
            if len(ln) < 4 or len(ln) > 50:
                continue
            if "@" in ln or "www" in ln.lower():
                continue
            if re.search(r"\d", ln):
                continue
            norm = normalize_company_name(ln)
            if len(norm) >= 4 and norm not in normalized_existing:
                cands.append(ln)
                break

    out = []
    seen = set()
    for c in cands:
        key = normalize_company_name(c)
        if key and key not in seen:
            seen.add(key)
            out.append(c)
    return out[:8]


def search_gouv_company_ranked(
    company_raw: str,
    email: str,
    website: str,
    city: str,
    ocr_text: str,
    address_hint: str = "",
    postal_code_hint: str = "",
) -> Dict[str, str]:
    candidates = build_company_candidates(company_raw, email, website, city, ocr_text)
    if not candidates:
        return {}

    brand_email = normalize_company_name(brand_from_domain(domain_from_email(email)))
    brand_web = normalize_company_name(brand_from_domain(domain_from_url(website)))
    city_n = normalize_text(city)
    addr_n = normalize_text(address_hint)
    cp_n = normalize_cp(postal_code_hint)

    best_score = -999
    best = {}

    for q in candidates:
        try:
            cache_key = f"search_gouv_company_ranked|{q}|{city}"
            if cache_key in GOUV_CACHE:
                results = GOUV_CACHE[cache_key].get("_results") or []
            else:
                STATS["gouv_calls"] += 1
                url = f"https://recherche-entreprises.api.gouv.fr/search?q={requests.utils.quote(q)}&page=1&per_page=8"
                r = requests.get(url, headers={"accept": "application/json"}, timeout=20)
                r.raise_for_status()
                j = r.json()
                results = j.get("results") or []
                GOUV_CACHE[cache_key] = {"_results": results}
        except Exception:
            continue

        for e in results:
            siege = e.get("siege") or {}
            denom = (e.get("nom_raison_sociale") or e.get("denomination") or "").strip()
            denom_n = normalize_company_name(denom)
            result_city = (siege.get("libelle_commune") or "").strip()
            result_city_n = normalize_text(result_city)
            result_addr = (siege.get("adresse") or siege.get("libelle_voie") or "").strip()
            result_addr_n = normalize_text(result_addr)
            result_cp = normalize_cp(siege.get("code_postal") or "")

            score = 0

            for token in [brand_email, brand_web, normalize_company_name(company_raw), normalize_company_name(q)]:
                if token and token in denom_n:
                    score += 20

            if city_n and result_city_n:
                if city_n == result_city_n:
                    score += 30
                elif city_n in result_city_n or result_city_n in city_n:
                    score += 15

            if cp_n and result_cp and cp_n == result_cp:
                score += 20

            if addr_n and result_addr_n:
                common = 0
                for tok in addr_n.split():
                    if len(tok) >= 3 and tok in result_addr_n:
                        common += 1
                score += min(common * 5, 20)

            if brand_email and brand_email in denom_n:
                score += 20
            if brand_web and brand_web in denom_n:
                score += 20

            if score > best_score:
                best_score = score

                dirigeant = ""
                arr = e.get("dirigeants") or e.get("representants") or []
                if arr:
                    first = arr[0]
                    if isinstance(first, str):
                        dirigeant = first
                    elif isinstance(first, dict):
                        p = first.get("personne") or first
                        prenom = p.get("prenom") or p.get("prenoms") or ""
                        nom = p.get("nom") or p.get("nom_usage") or ""
                        dirigeant = f"{prenom} {nom}".strip()

                naf = (e.get("activite_principale") or e.get("naf") or "").replace(".", "").replace(" ", "").upper()

                best = {
                    "name": denom,
                    "siret": validate_siret(siege.get("siret") or ""),
                    "naf": naf,
                    "address": result_addr,
                    "postal_code": result_cp,
                    "city": result_city,
                    "dirigeant": dirigeant,
                }

    return best


def search_gouv_by_address_hint(address: str, postal_code: str, city: str, website: str, email: str) -> Dict[str, str]:
    parts = []

    if address:
        parts.append(address)
    if postal_code:
        parts.append(postal_code)
    if city:
        parts.append(city)

    dom_web = brand_from_domain(domain_from_url(website))
    dom_email = brand_from_domain(domain_from_email(email))

    if dom_web:
        parts.append(dom_web)
    elif dom_email:
        parts.append(dom_email)

    q = " ".join([p for p in parts if p]).strip()
    if not q:
        return {}

    cache_key = f"search_gouv_by_address_hint|{q}"
    if cache_key in GOUV_CACHE:
        return dict(GOUV_CACHE[cache_key])

    try:
        STATS["gouv_calls"] += 1
        url = f"https://recherche-entreprises.api.gouv.fr/search?q={requests.utils.quote(q)}&page=1&per_page=6"
        r = requests.get(url, headers={"accept": "application/json"}, timeout=20)
        r.raise_for_status()
        j = r.json()
        results = j.get("results") or []
        if not results:
            GOUV_CACHE[cache_key] = {}
            return {}

        city_n = normalize_text(city)
        best_score = -999
        best = {}

        for e in results:
            siege = e.get("siege") or {}
            result_city = (siege.get("libelle_commune") or "").strip()
            result_city_n = normalize_text(result_city)
            denom = (e.get("nom_raison_sociale") or e.get("denomination") or "").strip()
            denom_n = normalize_company_name(denom)

            score = 0
            if city_n and result_city_n:
                if city_n == result_city_n:
                    score += 30
                elif city_n in result_city_n or result_city_n in city_n:
                    score += 15

            for tok in [dom_web, dom_email]:
                tok_n = normalize_company_name(tok)
                if tok_n and tok_n in denom_n:
                    score += 25

            dirigeant = ""
            arr = e.get("dirigeants") or e.get("representants") or []
            if arr:
                first = arr[0]
                if isinstance(first, str):
                    dirigeant = first
                elif isinstance(first, dict):
                    p = first.get("personne") or first
                    prenom = p.get("prenom") or p.get("prenoms") or ""
                    nom = p.get("nom") or p.get("nom_usage") or ""
                    dirigeant = f"{prenom} {nom}".strip()

            naf = (e.get("activite_principale") or e.get("naf") or "").replace(".", "").replace(" ", "").upper()

            if score > best_score:
                best_score = score
                best = {
                    "name": denom,
                    "siret": validate_siret(siege.get("siret") or ""),
                    "naf": naf,
                    "address": (siege.get("adresse") or siege.get("libelle_voie") or ""),
                    "postal_code": normalize_cp(siege.get("code_postal") or ""),
                    "city": result_city,
                    "dirigeant": dirigeant,
                }

        GOUV_CACHE[cache_key] = best
        return dict(best)
    except Exception:
        return {}


def places_enrich(name: str, city: str = "") -> Dict[str, str]:
    if not GOOGLE_PLACES_API_KEY or not name:
        return {}

    cache_key = f"{name}|{city}"
    if cache_key in PLACES_CACHE:
        return dict(PLACES_CACHE[cache_key])

    try:
        STATS["places_calls"] += 1
        query = name if not city else f"{name} {city}".strip()
        r1 = requests.post(
            "https://places.googleapis.com/v1/places:searchText",
            headers={
                "Content-Type": "application/json",
                "X-Goog-Api-Key": GOOGLE_PLACES_API_KEY,
                "X-Goog-FieldMask": "places.id,places.websiteUri",
            },
            json={
                "textQuery": query,
                "maxResultCount": 1,
                "languageCode": "fr",
                "regionCode": "FR",
            },
            timeout=20
        )
        r1.raise_for_status()
        j1 = r1.json()
        p = (j1.get("places") or [None])[0]
        if not p or not p.get("id"):
            PLACES_CACHE[cache_key] = {}
            return {}
        pid = p["id"]
        website = p.get("websiteUri") or ""

        r2 = requests.get(
            f"https://places.googleapis.com/v1/places/{pid}",
            headers={
                "X-Goog-Api-Key": GOOGLE_PLACES_API_KEY,
                "X-Goog-FieldMask": "nationalPhoneNumber,internationalPhoneNumber,websiteUri,formattedAddress",
            },
            timeout=20
        )
        r2.raise_for_status()
        j2 = r2.json()
        phone = j2.get("nationalPhoneNumber") or j2.get("internationalPhoneNumber") or ""
        website2 = j2.get("websiteUri") or website
        addr = (j2.get("formattedAddress") or "").strip()

        out = {"phone": normalize_phone(phone), "website": (website2 or "").strip(), "address": addr}
        PLACES_CACHE[cache_key] = out
        return dict(out)
    except Exception:
        return {}


def scrape_email_from_site(url: str) -> str:
    if not url:
        return ""

    domain = domain_from_url(url)
    if domain in EMAIL_SCRAPE_CACHE:
        cached = EMAIL_SCRAPE_CACHE[domain]
        return "" if cached == "none" else cached

    candidates = []
    base = url.rstrip("/")
    if base:
        candidates.append(base + "/contact")
        candidates.append(base)
    else:
        candidates.append(url)

    for attempt_url in candidates:
        try:
            STATS["scrape_calls"] += 1
            r = requests.get(attempt_url, headers={"user-agent": "Mozilla/5.0 (ProspectionBot)"}, timeout=10)
            if r.status_code >= 300:
                continue
            html = r.text[:250000]
            matches = re.findall(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", html)
            if not matches:
                continue

            valid = []
            for m in matches:
                em = clean_email(m)
                if not is_valid_email(em):
                    continue
                if re.search(r"no-?reply", em, re.I):
                    continue
                valid.append(em)

            if not valid:
                continue

            if domain:
                same_domain = [em for em in valid if domain_from_email(em) == domain]
                if same_domain:
                    EMAIL_SCRAPE_CACHE[domain] = same_domain[0]
                    return same_domain[0]

            EMAIL_SCRAPE_CACHE[domain] = valid[0]
            return valid[0]
        except Exception:
            continue

    if domain:
        EMAIL_SCRAPE_CACHE[domain] = "none"
    return ""


# ============================================================
# DOUBLE PASSE CARTE DE VISITE
# ============================================================

def extract_card_pipeline_fields(img_bytes: bytes) -> Dict[str, str]:
    if not img_bytes:
        return {}

    cache_key = image_hash_bytes(img_bytes)
    if cache_key in CARD_PIPELINE_CACHE:
        return dict(CARD_PIPELINE_CACHE[cache_key])

    # PASSE 1 — GEMINI STRUCTURÉ
    vis = gemini_extract_business_card(img_bytes) if GEMINI_API_KEY else {
        "name": "", "company": "", "title": "", "email": "", "phone": "",
        "website": "", "postal_code": "", "city": "", "address": ""
    }

    # PASSE 2 — EASYOCR + TESSERACT + REGEX
    light_ocr = bool(vis.get("name") or vis.get("company") or vis.get("email") or vis.get("phone") or vis.get("website"))
    ocr_txt = ocr_image_bytes(img_bytes, light=light_ocr)
    lines = clean_ocr_lines(ocr_txt)

    email_ocr = best_email("\n".join(lines))
    urls_ocr = extract_urls("\n".join(lines))
    phones_ocr = extract_all_phones("\n".join(lines))
    local_addr_ocr, local_cp_ocr, local_city_ocr = extract_local_address_from_text("\n".join(lines))

    local_website = (vis.get("website") or (urls_ocr[0] if urls_ocr else "") or "").strip()
    local_email = clean_email(vis.get("email") or email_ocr or "")
    if local_email and not is_valid_email(local_email):
        local_email = ""

    detected_company_line = detect_company_from_lines(
        lines,
        gemini_company=(vis.get("company") or "").strip(),
        website=local_website,
        email=local_email
    )

    person_name = (vis.get("name") or "").strip()
    if not person_name:
        person_name = detect_person_name_from_lines(lines, company_line=detected_company_line)

    fn, ln = split_human_name(person_name)

    local_postal_code = normalize_cp((vis.get("postal_code") or local_cp_ocr or "").strip())
    local_city = (vis.get("city") or local_city_ocr or "").strip()
    local_address = clean_address_candidate((vis.get("address") or local_addr_ocr or "").strip())

    phones_all = []
    if vis.get("phone"):
        phones_all.append(normalize_phone(vis.get("phone")))
    phones_all.extend(phones_ocr)

    seen = set()
    dedup_phones = []
    for p in phones_all:
        p = normalize_phone(p)
        if p and p not in seen:
            seen.add(p)
            dedup_phones.append(p)

    phone, phone2 = split_phones_by_priority(dedup_phones)

    company_guess = extract_best_company_hint(
        lines=lines,
        gemini_company=(vis.get("company") or "").strip(),
        website=local_website,
        email=local_email
    )

    gouv = search_gouv_company_ranked(
        company_raw=company_guess,
        email=local_email,
        website=local_website,
        city=local_city,
        ocr_text="\n".join(lines),
        address_hint=local_address,
        postal_code_hint=local_postal_code
    )

    if not gouv:
        gouv = search_gouv_by_address_hint(
            address=local_address,
            postal_code=local_postal_code,
            city=local_city,
            website=local_website,
            email=local_email
        )

    if not gouv and company_guess:
        gouv = search_gouv_company(company_guess, local_city)

    place = {}
    place_query_name = (gouv.get("name") or company_guess or "").strip()
    if place_query_name:
        place = places_enrich(place_query_name, local_city) if local_city else places_enrich(place_query_name, "")
        place = place or {}

    final_name = choose_final_company_name(
        gouv_name=(gouv.get("name") or "").strip(),
        gemini_company=(vis.get("company") or "").strip(),
        detected_company_line=(detected_company_line or "").strip(),
        person_name=(person_name or "").strip(),
        website=local_website or (place.get("website") or ""),
        email=local_email,
    )

    final_address = clean_address_candidate(
        (local_address or gouv.get("address") or place.get("address") or "").strip()
    )
    final_postal_code = normalize_cp(local_postal_code or gouv.get("postal_code") or "")
    final_city = (local_city or gouv.get("city") or "").strip()
    final_website = (local_website or place.get("website") or "").strip()
    final_email = local_email or (scrape_email_from_site(final_website) if final_website else "")
    final_email = clean_email(final_email) if is_valid_email(final_email) else ""

    if not phone and place.get("phone"):
        phone = normalize_phone(place.get("phone") or "")

    out = {
        "name": final_name,
        "address": final_address,
        "postal_code": final_postal_code,
        "city": final_city,
        "siret": validate_siret((gouv.get("siret") or "").strip()),
        "naf": (gouv.get("naf") or "").strip(),
        "activity_summary": enrich_activity_summary(
            name=final_name,
            naf=(gouv.get("naf") or "").strip(),
            website=final_website,
            company_hint=(company_guess or detected_company_line or "")
        ),
        "dirigeant": (gouv.get("dirigeant") or "").strip(),
        "interlocuteur": (person_name or "").strip(),
        "contact_firstname": fn,
        "contact_lastname": ln,
        "phone": phone,
        "phone2": phone2,
        "email": final_email,
        "website": final_website,
    }

    if out["phone"] and out["phone2"] and out["phone"] == out["phone2"]:
        out["phone2"] = ""

    out = ensure_contact_fallback(out)
    out = normalize_row_address(out)
    CARD_PIPELINE_CACHE[cache_key] = dict(out)
    return dict(out)


def is_linked_prospect_card(card: Dict[str, Any]) -> bool:
    if not isinstance(card, dict):
        return False
    source = (card.get("source") or "").strip().lower()
    source_mode = (card.get("source_mode") or "").strip().lower()
    comment = (card.get("comment") or "").strip().lower()

    if source == "prospect_mode":
        return True
    if source_mode == "company_form":
        return True
    if comment == "card_from_prospect_mode":
        return True
    if card.get("linked_prospect_key"):
        return True
    if card.get("linked_prospect_name"):
        return True
    return False


def merge_card_into_prospect_row(row: Dict[str, Any], img_bytes: bytes) -> Dict[str, Any]:
    if not img_bytes:
        return row

    card = extract_card_pipeline_fields(img_bytes)
    if not card:
        return row

    merged = dict(row)

    row_score = record_score(row)
    card_score = record_score(card)
    card_is_globally_better = card_score > row_score

    # Carte prioritaire sur API / Places pour les infos visibles,
    # mais sans écraser brutalement un prospect déjà meilleur.
    for fld in [
        "interlocuteur",
        "contact_firstname",
        "contact_lastname",
        "email",
        "website",
        "address",
        "postal_code",
        "city",
    ]:
        card_val = str(card.get(fld, "") or "").strip()
        row_val = str(merged.get(fld, "") or "").strip()
        if not card_val:
            continue

        should_replace = False
        if not row_val:
            should_replace = True
        elif fld == "email" and is_valid_email(card_val) and not is_valid_email(row_val):
            should_replace = True
        elif fld == "website" and len(card_val) > len(row_val):
            should_replace = True
        elif fld in ("address", "interlocuteur", "contact_firstname", "contact_lastname") and len(card_val) > len(row_val):
            should_replace = True
        elif fld in ("postal_code", "city") and not merged.get("address") and card_is_globally_better:
            should_replace = True

        if should_replace:
            merged[fld] = card.get(fld, "")

    if str(card.get("name", "") or "").strip():
        current_name = str(merged.get("name", "") or "").strip()
        card_name = str(card.get("name", "") or "").strip()
        if (not current_name) or (len(card_name) > len(current_name) and card_is_globally_better):
            merged["name"] = card.get("name", "")

    if not str(merged.get("siret", "") or "").strip() and str(card.get("siret", "") or "").strip():
        merged["siret"] = card.get("siret", "")
    if not str(merged.get("naf", "") or "").strip() and str(card.get("naf", "") or "").strip():
        merged["naf"] = card.get("naf", "")

    if not str(merged.get("dirigeant", "") or "").strip() and str(card.get("dirigeant", "") or "").strip():
        merged["dirigeant"] = card.get("dirigeant", "")

    card_phones = []
    for fld in ["phone", "phone2"]:
        p = normalize_phone(card.get(fld, "") or "")
        if p and p not in card_phones:
            card_phones.append(p)

    existing_phones = []
    for fld in ["phone", "phone2"]:
        p = normalize_phone(merged.get(fld, "") or "")
        if p and p not in existing_phones:
            existing_phones.append(p)

    if card_phones:
        phone, phone2 = split_phones_by_priority(card_phones + existing_phones)
        merged["phone"] = normalize_phone(phone)
        merged["phone2"] = normalize_phone(phone2)

    if merged.get("phone") and merged.get("phone2") and merged["phone"] == merged["phone2"]:
        merged["phone2"] = ""

    if not merged.get("activity_summary"):
        merged["activity_summary"] = enrich_activity_summary(
            name=merged.get("name", ""),
            naf=merged.get("naf", ""),
            website=merged.get("website", ""),
            company_hint=card.get("name", "") or merged.get("name", "")
        )

    merged = ensure_contact_fallback(merged)
    merged = normalize_row_address(merged)
    merged = uppercase_business_fields(merged)

    log_fusion("merge_card_into_prospect_row", {
        "prospect_name": row.get("name", ""),
        "card_name": card.get("name", ""),
        "card_interlocuteur": card.get("interlocuteur", ""),
        "card_email": card.get("email", ""),
    })

    return merged


def unify_from_prospect_with_card(p: Dict[str, Any]) -> Dict[str, Any]:
    row = unify_from_prospect(p)

    card_file_id = (p.get("card_photo_file_id") or "").strip()
    if not card_file_id:
        STATS["linked_cards_missing"] += 1
        return row

    try:
        img = tg_download_file(card_file_id)
    except Exception:
        img = b""

    if not img:
        STATS["linked_cards_missing"] += 1
        return row

    STATS["linked_cards_processed"] += 1
    return merge_card_into_prospect_row(row, img)


# ============================================================
# DÉDUPLICATION AVANCÉE
# ============================================================

def domain_match_value(row: Dict[str, Any]) -> str:
    website_dom = domain_from_url(row.get("website", "") or "")
    email_dom = domain_from_email(row.get("email", "") or "")
    return website_dom or email_dom


def row_identity_signals(row: Dict[str, Any]) -> Dict[str, str]:
    return {
        "siret": validate_siret(row.get("siret", "") or ""),
        "name": normalize_company_name(row.get("name", "") or ""),
        "city": normalize_text(row.get("city", "") or ""),
        "postal_code": normalize_cp(row.get("postal_code", "") or ""),
        "phone": normalize_phone(row.get("phone", "") or ""),
        "phone2": normalize_phone(row.get("phone2", "") or ""),
        "email": clean_email(row.get("email", "") or ""),
        "domain": domain_match_value(row),
        "website_domain": domain_from_url(row.get("website", "") or ""),
        "interlocuteur": normalize_text(row.get("interlocuteur", "") or ""),
        "address": normalize_text(row.get("address", "") or ""),
        "agency": normalize_text(row.get("agency", "") or ""),
        "initials": normalize_text(row.get("initials", "") or ""),
    }


def record_score(d: Dict[str, Any]) -> int:
    score = 0

    weights = {
        "name": 5,
        "address": 3,
        "postal_code": 2,
        "city": 2,
        "siret": 8,
        "naf": 2,
        "activity_summary": 2,
        "dirigeant": 2,
        "interlocuteur": 4,
        "contact_firstname": 2,
        "contact_lastname": 2,
        "phone": 3,
        "phone2": 2,
        "email": 4,
        "website": 2,
        "resume": 3,
        "commande": 3,
    }

    for k, w in weights.items():
        v = str(d.get(k, "") or "").strip()
        if not v:
            continue

        if k == "siret":
            if validate_siret(v):
                score += w
            continue

        if k in ("email",):
            if is_valid_email(v):
                score += w
            continue

        if k in ("phone", "phone2"):
            if normalize_phone(v):
                score += w
            continue

        score += w

    if validate_siret(d.get("siret", "") or ""):
        score += 4

    if str(d.get("resume", "") or "").strip() and str(d.get("commande", "") or "").strip():
        score += 2

    return score


def row_similarity_score(a: Dict[str, Any], b: Dict[str, Any]) -> int:
    sa = row_identity_signals(a)
    sb = row_identity_signals(b)
    score = 0

    if sa["siret"] and sb["siret"] and sa["siret"] == sb["siret"]:
        score += 100

    if sa["name"] and sb["name"]:
        if sa["name"] == sb["name"]:
            score += 30
        elif sa["name"] in sb["name"] or sb["name"] in sa["name"]:
            score += 18

    if sa["city"] and sb["city"]:
        if sa["city"] == sb["city"]:
            score += 12
        elif sa["city"] in sb["city"] or sb["city"] in sa["city"]:
            score += 6

    if sa["postal_code"] and sb["postal_code"] and sa["postal_code"] == sb["postal_code"]:
        score += 10

    phones_a = {x for x in [sa["phone"], sa["phone2"]] if x}
    phones_b = {x for x in [sb["phone"], sb["phone2"]] if x}
    if phones_a and phones_b:
        inter = phones_a & phones_b
        if inter:
            score += 20

    if sa["email"] and sb["email"] and sa["email"] == sb["email"]:
        score += 24

    if sa["domain"] and sb["domain"] and sa["domain"] == sb["domain"]:
        score += 12

    if sa["website_domain"] and sb["website_domain"] and sa["website_domain"] == sb["website_domain"]:
        score += 10

    if sa["interlocuteur"] and sb["interlocuteur"]:
        if sa["interlocuteur"] == sb["interlocuteur"]:
            score += 10
        elif sa["interlocuteur"] in sb["interlocuteur"] or sb["interlocuteur"] in sa["interlocuteur"]:
            score += 5

    if sa["address"] and sb["address"]:
        common = 0
        toks_a = [t for t in sa["address"].split() if len(t) >= 3]
        for tok in toks_a:
            if tok in sb["address"]:
                common += 1
        score += min(common * 2, 8)

    # proximité carte/prospect / même personne même jour
    if sa["agency"] and sb["agency"] and sa["agency"] == sb["agency"]:
        score += 2
    if sa["initials"] and sb["initials"] and sa["initials"] == sb["initials"]:
        score += 2

    return score


def dedupe_anchor_key(r: Dict[str, Any]) -> str:
    sig = row_identity_signals(r)

    if sig["siret"]:
        return f"SIRET|{sig['siret']}"

    if sig["name"] and sig["city"] and sig["postal_code"]:
        return f"NCP|{sig['name']}|{sig['city']}|{sig['postal_code']}"

    if sig["name"] and sig["city"]:
        return f"NC|{sig['name']}|{sig['city']}"

    if sig["email"]:
        return f"EMAIL|{sig['email']}"

    phones = sorted([x for x in [sig["phone"], sig["phone2"]] if x])
    if sig["name"] and phones:
        return f"NP|{sig['name']}|{'|'.join(phones)}"

    dom = sig["domain"]
    if sig["name"] and dom:
        return f"ND|{sig['name']}|{dom}"

    return ""


def merge_rows_keep_best(primary: Dict[str, Any], secondary: Dict[str, Any]) -> Dict[str, Any]:
    """
    Garde la ligne la plus complète, sans écraser les champs terrain forts.
    Priorité spéciale :
    - resume / commande : ne jamais écraser si déjà présents dans primary
    - dirigeant : ne pas écraser un dirigeant existant par vide
    """
    merged = dict(primary)

    protected_if_present = {"resume", "commande"}
    for k, v2 in secondary.items():
        v1 = merged.get(k, "")
        s1 = str(v1 or "").strip()
        s2 = str(v2 or "").strip()

        if k in protected_if_present and s1:
            continue

        if not s1 and s2:
            merged[k] = v2
            continue

        if s1 and s2:
            if k == "activity_summary":
                if len(s2) > len(s1):
                    merged[k] = v2
            elif k == "address":
                if len(s2) > len(s1):
                    merged[k] = v2
            elif k in ("phone", "phone2"):
                # éviter doublons, mais permettre meilleure affectation
                if not normalize_phone(s1) and normalize_phone(s2):
                    merged[k] = v2
            elif k == "website":
                # préfère URL réelle à vide ou à domaine reconstruit très court
                if len(s2) > len(s1):
                    merged[k] = v2
            elif k == "email":
                if is_valid_email(s2) and not is_valid_email(s1):
                    merged[k] = v2
            elif k == "siret":
                if validate_siret(s2) and not validate_siret(s1):
                    merged[k] = v2
            elif k == "naf":
                if s2 and not s1:
                    merged[k] = v2
            elif k == "interlocuteur":
                if len(s2) > len(s1):
                    merged[k] = v2
            elif k == "contact_firstname":
                if len(s2) > len(s1):
                    merged[k] = v2
            elif k == "contact_lastname":
                if len(s2) > len(s1):
                    merged[k] = v2

    merged = ensure_contact_fallback(merged)
    merged = normalize_row_address(merged)

    p1 = normalize_phone(merged.get("phone", "") or "")
    p2 = normalize_phone(merged.get("phone2", "") or "")
    if p1 and p2 and p1 == p2:
        merged["phone2"] = ""

    return merged


def dedupe_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not rows:
        return []

    buckets: Dict[str, List[Dict[str, Any]]] = {}
    loose_rows: List[Dict[str, Any]] = []

    for r in rows:
        key = dedupe_anchor_key(r)
        if key:
            buckets.setdefault(key, []).append(r)
        else:
            loose_rows.append(r)

    merged_groups: List[Dict[str, Any]] = []

    for anchor, group in buckets.items():
        STATS["dedupe_groups"] += 1
        consumed = [False] * len(group)

        for i in range(len(group)):
            if consumed[i]:
                continue
            base = dict(group[i])
            consumed[i] = True

            for j in range(i + 1, len(group)):
                if consumed[j]:
                    continue
                sim = row_similarity_score(base, group[j])
                if sim >= 28:
                    before = record_score(base)
                    base = merge_rows_keep_best(base, group[j])
                    after = record_score(base)
                    consumed[j] = True
                    STATS["dedupe_rows_merged"] += 1
                    log_fusion("dedupe_merge", {
                        "anchor": anchor,
                        "similarity": sim,
                        "score_before": before,
                        "score_after": after,
                        "name": base.get("name", ""),
                        "city": base.get("city", ""),
                    })

            merged_groups.append(base)

    # deuxième passe légère pour les lignes sans clé forte
    out = list(merged_groups)
    for r in loose_rows:
        attached = False
        for idx, ex in enumerate(out):
            sim = row_similarity_score(r, ex)
            if sim >= 34:
                before = record_score(out[idx])
                out[idx] = merge_rows_keep_best(ex, r)
                after = record_score(out[idx])
                STATS["dedupe_rows_merged"] += 1
                log_fusion("dedupe_loose_attach", {
                    "similarity": sim,
                    "score_before": before,
                    "score_after": after,
                    "name": out[idx].get("name", ""),
                    "city": out[idx].get("city", ""),
                })
                attached = True
                break
        if not attached:
            out.append(r)

    return out


# ============================================================
# UNIFY
# ============================================================

def unify_from_prospect(p: Dict[str, Any]) -> Dict[str, Any]:
    website = get_first_valid(p, "website", "site_web", "websiteUri")
    naf = get_first_valid(p, "naf", "code_naf", "activite_principale")
    name = get_first_valid(p, "name", "company", "nom", "denomination")
    activity_summary = get_first_valid(p, "activity_summary", "activity", "activity_label")

    if not activity_summary and (name or naf or website):
        activity_summary = enrich_activity_summary(name=name, naf=naf, website=website, company_hint=name)

    raw_email = get_first_valid(p, "email", "contact_email", "mail")
    email = clean_email(raw_email) if is_valid_email(raw_email) else ""

    phones = []
    for k in ["phone", "phone2", "mobile", "telephone", "tel"]:
        v = normalize_phone(p.get(k, "") or "")
        if v and v not in phones:
            phones.append(v)
    phone, phone2 = split_phones_by_priority(phones)

    d = {
        "date": get_first_valid(p, "date") or "",
        "agency": (get_first_valid(p, "agency") or "").upper(),
        "initials": (get_first_valid(p, "initials", "user_initials") or "").upper(),
        "name": name,
        "address": get_first_valid(p, "address", "adresse", "formattedAddress"),
        "postal_code": get_first_valid(p, "postal_code", "code_postal", "postalCode"),
        "city": get_first_valid(p, "city", "ville", "commune"),
        "siret": validate_siret(get_first_valid(p, "siret", "siretSiege", "siret_siege")),
        "naf": naf,
        "activity_summary": activity_summary,
        "dirigeant": get_first_valid(p, "dirigeant", "representant_legal", "nom_dirigeant"),
        "interlocuteur": get_first_valid(p, "interlocuteur", "contact_name", "person_name"),
        "contact_firstname": get_first_valid(p, "contact_firstname", "firstname", "first_name"),
        "contact_lastname": get_first_valid(p, "contact_lastname", "lastname", "last_name"),
        "phone": phone,
        "phone2": phone2,
        "email": email,
        "website": website,
        "resume": get_first_valid(p, "resume", "meeting", "comment"),
        "commande": get_first_valid(p, "commande", "order"),
    }

    if d["phone"] and d["phone2"] and d["phone"] == d["phone2"]:
        d["phone2"] = ""

    d = ensure_contact_fallback(d)
    d = normalize_row_address(d)
    d = uppercase_business_fields(d)
    return d


def unify_from_card(date: str, agency: str, initials: str, comment: str, img_bytes: bytes) -> Dict[str, Any]:
    base = {
        "date": date, "agency": agency, "initials": initials,
        "name": "", "address": "", "postal_code": "", "city": "",
        "siret": "", "naf": "", "activity_summary": "", "dirigeant": "",
        "interlocuteur": "", "contact_firstname": "", "contact_lastname": "",
        "phone": "", "phone2": "", "email": "", "website": "",
        "resume": (comment or "").strip(),
        "commande": "",
    }
    if not img_bytes:
        return base

    fields = extract_card_pipeline_fields(img_bytes)
    if not fields:
        return base

    row = dict(base)
    row.update({
        "name": fields.get("name", ""),
        "address": fields.get("address", ""),
        "postal_code": fields.get("postal_code", ""),
        "city": fields.get("city", ""),
        "siret": validate_siret(fields.get("siret", "")),
        "naf": fields.get("naf", ""),
        "activity_summary": fields.get("activity_summary", ""),
        "dirigeant": fields.get("dirigeant", ""),
        "interlocuteur": fields.get("interlocuteur", ""),
        "contact_firstname": fields.get("contact_firstname", ""),
        "contact_lastname": fields.get("contact_lastname", ""),
        "phone": normalize_phone(fields.get("phone", "")),
        "phone2": normalize_phone(fields.get("phone2", "")),
        "email": clean_email(fields.get("email", "")) if is_valid_email(fields.get("email", "")) else "",
        "website": fields.get("website", ""),
    })

    if row["phone"] and row["phone2"] and row["phone"] == row["phone2"]:
        row["phone2"] = ""

    row = ensure_contact_fallback(row)
    row = normalize_row_address(row)
    row = uppercase_business_fields(row)
    return row


def unify_from_photo(date: str, agency: str, initials: str, city_hint: str, comment: str, img_bytes: bytes) -> Dict[str, Any]:
    base = {
        "date": date, "agency": agency, "initials": initials,
        "name": "", "address": "", "postal_code": "", "city": (city_hint or "").strip(),
        "siret": "", "naf": "", "activity_summary": "", "dirigeant": "",
        "interlocuteur": "", "contact_firstname": "", "contact_lastname": "",
        "phone": "", "phone2": "", "email": "", "website": "",
        "resume": (comment or "").strip(),
        "commande": "",
    }
    if not img_bytes:
        return base

    cache_key = image_hash_bytes(img_bytes)
    if cache_key in PHOTO_PIPELINE_CACHE:
        row = dict(base)
        row.update(dict(PHOTO_PIPELINE_CACHE[cache_key]))
        row = ensure_contact_fallback(row)
        row = normalize_row_address(row)
        row = uppercase_business_fields(row)
        return row

    vis = gemini_extract_facade_logo(img_bytes) if GEMINI_API_KEY else {"company": "", "city": ""}
    ocr_txt = ocr_image_bytes(img_bytes, light=not bool(vis.get("company")))
    lines = clean_ocr_lines(ocr_txt)

    company_raw = (vis.get("company") or "").strip()
    city = (vis.get("city") or "").strip() or (city_hint or "").strip()

    email_ocr = best_email("\n".join(lines))
    urls_ocr = extract_urls("\n".join(lines))
    phones_ocr = extract_all_phones("\n".join(lines))
    local_addr_ocr, local_cp_ocr, local_city_ocr = extract_local_address_from_text("\n".join(lines))

    website = (urls_ocr[0] if urls_ocr else "").strip()
    phone, phone2 = split_phones_by_priority(phones_ocr)

    if not city:
        city = local_city_ocr

    detected_company_line = detect_company_from_lines(
        lines=lines,
        gemini_company=company_raw,
        website=website,
        email=email_ocr
    )

    company_guess = extract_best_company_hint(
        lines=lines,
        gemini_company=company_raw,
        website=website,
        email=email_ocr
    )

    if not company_guess:
        return base

    gouv = search_gouv_company_ranked(
        company_raw=company_guess,
        email=email_ocr,
        website=website,
        city=city,
        ocr_text="\n".join(lines),
        address_hint=local_addr_ocr,
        postal_code_hint=local_cp_ocr
    )
    if not gouv:
        gouv = search_gouv_company(company_guess, city)

    place = places_enrich(gouv.get("name") or company_guess, city) if GOOGLE_PLACES_API_KEY else {}
    final_website = (website or place.get("website") or "").strip()
    final_email = (email_ocr or "").strip().lower() or (scrape_email_from_site(final_website) if final_website else "")
    final_email = clean_email(final_email) if is_valid_email(final_email) else ""

    if not phone and place.get("phone"):
        phone = normalize_phone(place.get("phone") or "")

    final_name = choose_final_company_name(
        gouv_name=(gouv.get("name") or "").strip(),
        gemini_company=(company_raw or "").strip(),
        detected_company_line=(detected_company_line or "").strip(),
        person_name="",
        website=final_website,
        email=final_email,
    )

    row = dict(base)
    row.update({
        "name": final_name,
        "address": strip_postal_city_from_address(
            clean_address_candidate((local_addr_ocr or gouv.get("address") or place.get("address") or "").strip()),
            local_cp_ocr or (gouv.get("postal_code") or ""),
            local_city_ocr or (gouv.get("city") or city),
        ),
        "postal_code": (local_cp_ocr or gouv.get("postal_code") or "").strip(),
        "city": (local_city_ocr or gouv.get("city") or city).strip(),
        "siret": validate_siret((gouv.get("siret") or "").strip()),
        "naf": (gouv.get("naf") or "").strip(),
        "activity_summary": enrich_activity_summary(
            name=final_name,
            naf=(gouv.get("naf") or "").strip(),
            website=final_website,
            company_hint=company_raw
        ),
        "dirigeant": (gouv.get("dirigeant") or "").strip(),
        "phone": phone,
        "phone2": phone2,
        "website": final_website,
        "email": final_email,
    })

    if row["phone"] and row["phone2"] and row["phone"] == row["phone2"]:
        row["phone2"] = ""

    PHOTO_PIPELINE_CACHE[cache_key] = {
        k: row.get(k, "") for k in [
            "name", "address", "postal_code", "city", "siret", "naf", "activity_summary",
            "dirigeant", "phone", "phone2", "website", "email"
        ]
    }

    row = ensure_contact_fallback(row)
    row = normalize_row_address(row)
    row = uppercase_business_fields(row)
    return row


# ============================================================
# CARTES LIÉES / COMPAT HISTORIQUE / LOT -> PROSPECT
# ============================================================

def safe_get(d: Dict[str, Any], *keys: str) -> Any:
    for k in keys:
        if k in d and d.get(k) not in (None, ""):
            return d.get(k)
    return ""


def get_first_valid(d: Dict[str, Any], *keys: str) -> str:
    for k in keys:
        v = d.get(k)
        if v is None:
            continue
        if isinstance(v, str):
            if v.strip():
                return v.strip()
            continue
        if isinstance(v, (int, float)):
            return str(v).strip()
    return ""


def card_user_initials(card: Dict[str, Any]) -> str:
    return (safe_get(card, "user", "initials") or "").upper().strip()


def prospect_initials(prospect: Dict[str, Any]) -> str:
    return (prospect.get("initials") or "").upper().strip()


def card_link_key_candidates(card: Dict[str, Any]) -> List[str]:
    out = []

    for v in [
        card.get("linked_prospect_key", ""),
        card.get("card_kv_key", ""),
        card.get("prospect_key", ""),
    ]:
        v = safe_strip(v)
        if v and v not in out:
            out.append(v)

    linked_name = safe_strip(card.get("linked_prospect_name", ""))
    if linked_name:
        key = normalize_company_name(linked_name)
        if key and key not in out:
            out.append(key)

    gem = card.get("gemini") or {}
    g_company = safe_strip(gem.get("company", ""))
    if g_company:
        key = normalize_company_name(g_company)
        if key and key not in out:
            out.append(key)

    comment = safe_strip(card.get("comment", ""))
    if comment:
        low = normalize_company_name(comment)
        if low and len(low) >= 4 and low not in out:
            out.append(low)

    return out


def prospect_link_key_candidates(p: Dict[str, Any]) -> List[str]:
    out = []

    for v in [
        p.get("card_kv_key", ""),
        p.get("prospect_key", ""),
        p.get("linked_prospect_key", ""),
    ]:
        v = safe_strip(v)
        if v and v not in out:
            out.append(v)

    n = normalize_company_name(p.get("name", "") or "")
    if n and n not in out:
        out.append(n)

    return out


def card_is_same_scope(card: Dict[str, Any], prospect: Dict[str, Any]) -> bool:
    card_ag = (card.get("agency") or "").upper().strip()
    pros_ag = (prospect.get("agency") or "").upper().strip()

    card_ini = card_user_initials(card)
    pros_ini = prospect_initials(prospect)

    if card_ag and pros_ag and card_ag != pros_ag:
        return False

    if card_ini and pros_ini and card_ini != pros_ini:
        return False

    return True


def card_prospect_proximity_score(card: Dict[str, Any], prospect: Dict[str, Any]) -> int:
    score = 0

    if not card_is_same_scope(card, prospect):
        return -999

    ckeys = set(card_link_key_candidates(card))
    pkeys = set(prospect_link_key_candidates(prospect))
    if ckeys and pkeys and (ckeys & pkeys):
        score += 80

    linked_name = normalize_company_name(card.get("linked_prospect_name", "") or "")
    pname = normalize_company_name(prospect.get("name", "") or "")
    if linked_name and pname:
        if linked_name == pname:
            score += 35
        elif linked_name in pname or pname in linked_name:
            score += 20

    card_g = card.get("gemini") or {}
    card_company = normalize_company_name(card_g.get("company", "") or "")
    if card_company and pname:
        if card_company == pname:
            score += 24
        elif card_company in pname or pname in card_company:
            score += 12

    card_city = normalize_text(card_g.get("city", "") or card.get("city", "") or "")
    pros_city = normalize_text(prospect.get("city", "") or "")
    if card_city and pros_city:
        if card_city == pros_city:
            score += 8
        elif card_city in pros_city or pros_city in card_city:
            score += 4

    return score


def find_related_cards_for_prospect(prospect: Dict[str, Any], cards: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    matches: List[Tuple[int, Dict[str, Any]]] = []

    for c in cards:
        score = card_prospect_proximity_score(c, prospect)
        if score >= 18:
            matches.append((score, c))

    matches.sort(key=lambda x: x[0], reverse=True)

    out = []
    seen = set()
    for score, c in matches:
        key = f"{c.get('file_id','')}|{c.get('ts','')}|{c.get('linked_prospect_key','')}"
        if key in seen:
            continue
        seen.add(key)
        out.append(c)

    return out


def confidence_score_card_row(row: Dict[str, Any]) -> int:
    score = 0
    if safe_strip(row.get("name", "")):
        score += 25
    if safe_strip(row.get("city", "")):
        score += 10
    if normalize_cp(row.get("postal_code", "")):
        score += 8
    if safe_strip(row.get("address", "")):
        score += 8
    if validate_siret(row.get("siret", "")):
        score += 20
    if safe_strip(row.get("interlocuteur", "")):
        score += 10
    if is_valid_email(row.get("email", "")):
        score += 8
    if normalize_phone(row.get("phone", "")) or normalize_phone(row.get("phone2", "")):
        score += 6
    if safe_strip(row.get("website", "")):
        score += 5
    return score


def confidence_score_photo_row(row: Dict[str, Any]) -> int:
    score = 0
    if safe_strip(row.get("name", "")):
        score += 30
    if safe_strip(row.get("city", "")):
        score += 12
    if validate_siret(row.get("siret", "")):
        score += 20
    if safe_strip(row.get("address", "")):
        score += 10
    if is_valid_email(row.get("email", "")):
        score += 6
    if normalize_phone(row.get("phone", "")):
        score += 6
    if safe_strip(row.get("website", "")):
        score += 6
    return score


def unify_from_prospect_with_linked_cards(p: Dict[str, Any], cards: List[Dict[str, Any]]) -> Dict[str, Any]:
    row = unify_from_prospect(p)

    # 1. priorité à la carte directement portée par le prospect
    direct_card_file_id = safe_strip(p.get("card_photo_file_id", ""))
    if direct_card_file_id:
        try:
            img = tg_download_file(direct_card_file_id)
        except Exception:
            img = b""
        if img:
            STATS["linked_cards_processed"] += 1
            row = merge_card_into_prospect_row(row, img)
            log_fusion("prospect_direct_card_photo_file_id", {
                "prospect_name": p.get("name", ""),
                "agency": p.get("agency", ""),
                "initials": p.get("initials", ""),
            })

    # 2. fallback historique / KV / rapprochement
    related_cards = find_related_cards_for_prospect(p, cards)
    if related_cards:
        for c in related_cards:
            fid = c.get("file_id") or ""
            if not fid:
                continue
            try:
                img = tg_download_file(fid)
            except Exception:
                img = b""
            if not img:
                continue
            STATS["linked_cards_processed"] += 1
            row = merge_card_into_prospect_row(row, img)

        log_fusion("prospect_related_cards", {
            "prospect_name": p.get("name", ""),
            "cards_count": len(related_cards),
            "agency": p.get("agency", ""),
            "initials": p.get("initials", ""),
        })
    else:
        STATS["linked_cards_missing"] += 1

    # ne jamais écraser les champs terrain
    row["resume"] = p.get("resume", "") or row.get("resume", "")
    row["commande"] = p.get("commande", "") or row.get("commande", "")

    row = ensure_contact_fallback(row)
    row = normalize_row_address(row)
    row = uppercase_business_fields(row)
    return row


def standalone_cards(cards: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out = []
    for c in cards:
        if is_linked_prospect_card(c):
            continue
        out.append(c)
    return out


def promote_card_lot_to_prospect(card: Dict[str, Any], date: str) -> Optional[Dict[str, Any]]:
    fid = card.get("file_id") or ""
    if not fid:
        return None

    try:
        img = tg_download_file(fid)
    except Exception:
        img = b""
    if not img:
        return None

    agency = (card.get("agency") or "").upper().strip()
    initials = card_user_initials(card)
    comment = safe_strip(card.get("comment", ""))

    row = unify_from_card(date, agency, initials, comment, img)
    score = confidence_score_card_row(row)
    row["_confidence"] = score
    row["_source_type"] = "lot_card"

    if score < 30:
        return None

    STATS["lot_cards_promoted"] += 1
    log_fusion("promote_card_lot_to_prospect", {
        "name": row.get("name", ""),
        "city": row.get("city", ""),
        "confidence": score,
        "agency": agency,
        "initials": initials,
    })
    return row


def promote_photo_lot_to_prospect(photo: Dict[str, Any], date: str) -> Optional[Dict[str, Any]]:
    fid = photo.get("file_id") or ""
    if not fid:
        return None

    try:
        img = tg_download_file(fid)
    except Exception:
        img = b""
    if not img:
        return None

    agency = (photo.get("agency") or "").upper().strip()
    initials = (safe_get(photo, "user", "initials") or "").upper().strip()
    city_hint = photo.get("city") or ""
    comment = safe_strip(photo.get("comment") or photo.get("meeting") or "")

    row = unify_from_photo(date, agency, initials, city_hint, comment, img)
    score = confidence_score_photo_row(row)
    row["_confidence"] = score
    row["_source_type"] = "lot_photo"

    if score < 32:
        return None

    STATS["lot_photos_promoted"] += 1
    log_fusion("promote_photo_lot_to_prospect", {
        "name": row.get("name", ""),
        "city": row.get("city", ""),
        "confidence": score,
        "agency": agency,
        "initials": initials,
    })
    return row


def consolidate_rows_for_send_mode(
    date: str,
    prospects: List[Dict[str, Any]],
    photos: List[Dict[str, Any]],
    cards: List[Dict[str, Any]],
    agency_filter: str = "",
    initials_filter: str = "",
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    agency_filter = (agency_filter or "").upper().strip()
    initials_filter = (initials_filter or "").upper().strip()

    filtered_prospects = []
    for p in prospects:
        p_ag = (p.get("agency") or "").upper().strip()
        p_ini = prospect_initials(p)

        if agency_filter and p_ag != agency_filter:
            continue
        if initials_filter and p_ini != initials_filter:
            continue
        filtered_prospects.append(p)

    filtered_photos = []
    for ph in photos:
        ph_ag = (ph.get("agency") or "").upper().strip()
        ph_ini = (safe_get(ph, "user", "initials") or "").upper().strip()

        if agency_filter and ph_ag != agency_filter:
            continue
        if initials_filter and ph_ini != initials_filter:
            continue
        filtered_photos.append(ph)

    filtered_cards = []
    for ca in cards:
        ca_ag = (ca.get("agency") or "").upper().strip()
        ca_ini = card_user_initials(ca)

        if agency_filter and ca_ag != agency_filter:
            continue
        if initials_filter and ca_ini != initials_filter:
            continue
        filtered_cards.append(ca)

    # prospects enrichis par cartes liées
    for p in filtered_prospects:
        row = unify_from_prospect_with_linked_cards(p, filtered_cards)
        rows.append(row)

    # lots standalone cartes
    for c in standalone_cards(filtered_cards)[:MAX_OCR_IMAGES]:
        row = promote_card_lot_to_prospect(c, date)
        if row:
            rows.append(row)

    # lots façades
    for ph in filtered_photos[:MAX_PHOTO_IMAGES]:
        row = promote_photo_lot_to_prospect(ph, date)
        if row:
            rows.append(row)

    # nettoyage final
    out = []
    for r in rows:
        if not any(str(v or "").strip() for v in r.values()):
            continue

        r["date"] = r.get("date") or date
        r["agency"] = (r.get("agency") or agency_filter or "").upper()
        r["initials"] = (r.get("initials") or initials_filter or "").upper()

        r = ensure_contact_fallback(r)
        r = normalize_row_address(r)
        r = uppercase_business_fields(r)

        out.append(r)

    out = dedupe_rows(out)
    return out


# ============================================================
# EXCEL
# ============================================================

def to_row(d: Dict[str, Any]) -> List[Any]:
    return [d.get(h, "") for h in HEADERS]


def build_excel_one_sheet(date: str, rows: List[Dict[str, Any]], suffix: str) -> str:
    rows = dedupe_rows(rows)
    rows_sorted = sorted(rows, key=lambda x: (record_score(x), int(x.get("_confidence", 0))), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "DATA"
    ws.append(HEADERS)

    for r in rows_sorted:
        ws.append(to_row(r))

    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    autosize(ws)

    filename = os.path.join(OUT_DIR, f"PROSPECTION_{date}_{suffix}.xlsx")
    wb.save(filename)
    return filename


# ============================================================
# MEDIA ZIP
# ============================================================

def build_media_zip(
    date: str,
    agency: str,
    initials: str,
    photos: List[Dict[str, Any]],
    cards: List[Dict[str, Any]]
) -> Optional[Tuple[str, bytes]]:
    agency = (agency or "").upper().strip()
    initials = (initials or "").upper().strip()

    photos_u = [
        p for p in photos
        if (p.get("agency") or "").upper() == agency
        and (safe_get(p, "user", "initials") or "").upper().strip() == initials
    ]
    cards_u = [
        c for c in cards
        if (c.get("agency") or "").upper() == agency
        and card_user_initials(c) == initials
    ]

    if not photos_u and not cards_u:
        return None

    photos_u = photos_u[:MAX_PHOTO_IMAGES]
    cards_u = cards_u[:MAX_OCR_IMAGES]

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        lines = ["type;file;date;agency;initials;city;comment;geo_lat;geo_lon"]

        for i, p in enumerate(photos_u, start=1):
            fid = p.get("file_id") or ""
            try:
                img = tg_download_file(fid) if fid else b""
            except Exception:
                img = b""
            if not img:
                continue

            fname = f"photos/{date}_{agency}_{initials}_{i:02d}.jpg"
            z.writestr(fname, img)
            geo = p.get("geo") or {}
            comment = p.get("comment") or p.get("meeting") or ""
            lines.append(
                f"photo;{fname};{date};{agency};{initials};{(p.get('city') or '')};"
                f"{comment};{geo.get('lat') or ''};{geo.get('lon') or ''}"
            )

        for i, c in enumerate(cards_u, start=1):
            fid = c.get("file_id") or ""
            try:
                img = tg_download_file(fid) if fid else b""
            except Exception:
                img = b""
            if not img:
                continue

            fname = f"cards/{date}_{agency}_{initials}_{i:02d}.jpg"
            z.writestr(fname, img)
            lines.append(f"card;{fname};{date};{agency};{initials};;{(c.get('comment') or '')};;")

        z.writestr("index.csv", ("\n".join(lines)).encode("utf-8"))

    zip_name = f"MEDIA_{date}_{agency}_{initials}.zip"
    return zip_name, buf.getvalue()


# ============================================================
# BREVO EMAIL
# ============================================================

def brevo_send_email(
    to_email: str,
    subject: str,
    html: str,
    attachments: Optional[List[Tuple[str, bytes]]] = None
):
    if not BREVO_API_KEY:
        raise RuntimeError("BREVO_API_KEY missing")
    to_email = clean_email(to_email)
    if not is_valid_email(to_email):
        raise RuntimeError(f"Invalid recipient email: '{to_email}'")

    payload = {
        "sender": {"name": BREVO_SENDER_NAME, "email": BREVO_SENDER_EMAIL},
        "to": [{"email": to_email}],
        "subject": subject,
        "htmlContent": html,
    }

    if attachments:
        payload["attachment"] = []
        for filename, content in attachments:
            payload["attachment"].append({
                "name": filename,
                "content": base64.b64encode(content).decode("ascii")
            })

    r = requests.post(
        "https://api.brevo.com/v3/smtp/email",
        headers={
            "accept": "application/json",
            "api-key": BREVO_API_KEY,
            "content-type": "application/json"
        },
        data=json.dumps(payload),
        timeout=60
    )
    if r.status_code >= 300:
        raise RuntimeError(f"Brevo send failed {r.status_code}: {r.text[:400]}")
    STATS["emails_sent"] += 1


# ============================================================
# SEND MODES
# ============================================================

def send_individual_pack(
    date: str,
    agency: str,
    initials: str,
    prospects: List[Dict[str, Any]],
    photos: List[Dict[str, Any]],
    cards: List[Dict[str, Any]]
):
    to_email = email_for_initials(initials)
    if not to_email:
        print(f"[WARN] No email for initials={initials}, skip individual pack.")
        return

    initials = initials.upper().strip()
    agency = agency.upper().strip()

    rows = consolidate_rows_for_send_mode(
        date=date,
        prospects=prospects,
        photos=photos,
        cards=cards,
        agency_filter=agency,
        initials_filter=initials,
    )
    rows = [r for r in rows if any(str(v or "").strip() for v in r.values())]

    if not rows:
        print(f"[INFO] No rows for {agency}/{initials}, skip.")
        return

    xlsx = build_excel_one_sheet(date, rows, f"INDIV_{agency}_{initials}")

    attachments: List[Tuple[str, bytes]] = []
    with open(xlsx, "rb") as f:
        attachments.append((os.path.basename(xlsx), f.read()))

    media = build_media_zip(date, agency, initials, photos, cards)
    if media:
        attachments.append(media)

    subject = f"Prospection {date} — {agency}/{initials} (Excel + médias)"
    html = (
        f"<p>Bonjour,</p>"
        f"<p>Voici ton export de prospection du <b>{date}</b> pour <b>{agency}/{initials}</b>.</p>"
        f"<ul>"
        f"<li><b>Excel</b> consolidé, trié par complétude</li>"
        f"<li><b>ZIP médias</b> (photos + cartes)</li>"
        f"</ul>"
        f"<p>— Bot Prospection</p>"
    )

    brevo_send_email(to_email, subject, html, attachments=attachments)
    print(f"[OK] individual pack sent to {to_email} ({agency}/{initials})")


def send_agency_manager_pack(
    date: str,
    agency: str,
    prospects: List[Dict[str, Any]],
    photos: List[Dict[str, Any]],
    cards: List[Dict[str, Any]]
):
    agencies_cfg = ROUTING.get("agencies") or {}
    cfg = agencies_cfg.get(agency) or {}
    manager = (cfg.get("manager") or {})
    to_email = clean_email(manager.get("email") or "")
    if not is_valid_email(to_email):
        print(f"[WARN] No valid manager email for agency={agency} ({to_email})")
        return

    rows = consolidate_rows_for_send_mode(
        date=date,
        prospects=prospects,
        photos=photos,
        cards=cards,
        agency_filter=agency,
        initials_filter="",
    )
    rows = [r for r in rows if any(str(v or "").strip() for v in r.values())]

    if not rows:
        print(f"[INFO] No rows for agency={agency}, skip manager mail.")
        return

    xlsx = build_excel_one_sheet(date, rows, f"AGENCE_{agency}")

    with open(xlsx, "rb") as f:
        attachments = [(os.path.basename(xlsx), f.read())]

    html = (
        f"<p>Bonjour,</p>"
        f"<p>Voici le consolidé de prospection du <b>{date}</b> pour l’agence <b>{agency}</b>.</p>"
        f"<p>— Bot Prospection</p>"
    )

    brevo_send_email(
        to_email,
        f"Prospection {date} — Agence {agency} (consolidé)",
        html,
        attachments=attachments
    )
    print(f"[OK] agency manager pack sent to {to_email} (agency={agency})")


def send_admin_pack(
    date: str,
    prospects: List[Dict[str, Any]],
    photos: List[Dict[str, Any]],
    cards: List[Dict[str, Any]]
):
    admin = ROUTING.get("admin") or {}
    to_email = clean_email(admin.get("email") or "")
    if not is_valid_email(to_email):
        print(f"[WARN] No valid admin email configured ({to_email})")
        return

    rows = consolidate_rows_for_send_mode(
        date=date,
        prospects=prospects,
        photos=photos,
        cards=cards,
        agency_filter="",
        initials_filter="",
    )
    rows = [r for r in rows if any(str(v or "").strip() for v in r.values())]

    if not rows:
        print(f"[INFO] No rows for admin date={date}, skip.")
        return

    xlsx = build_excel_one_sheet(date, rows, "ADMIN_ALL")

    with open(xlsx, "rb") as f:
        attachments = [(os.path.basename(xlsx), f.read())]

    html = (
        f"<p>Bonjour,</p>"
        f"<p>Consolidé global de prospection du <b>{date}</b> pour toutes les agences.</p>"
        f"<p>— Bot Prospection</p>"
    )

    brevo_send_email(
        to_email,
        f"Prospection {date} — ADMIN (toutes agences)",
        html,
        attachments=attachments
    )
    print(f"[OK] admin pack sent to {to_email}")


# ============================================================
# MAIN
# ============================================================

def main():
    if not WORKER_BASE_URL or not EXPORT_TOKEN or not TELEGRAM_TOKEN:
        raise RuntimeError("Missing WORKER_BASE_URL / EXPORT_TOKEN / TELEGRAM_TOKEN")

    run_date = RUN_DATE if RUN_DATE else paris_ymd_fallback()

    print(f"🚀 export_and_mail.py — mode={SEND_MODE} date={run_date} agency={AGENCY} initials={INITIALS}")
    print(f"📦 OUT_DIR={OUT_DIR}")
    print(f"🔑 Gemini={'ON' if GEMINI_API_KEY else 'OFF'} | Places={'ON' if GOOGLE_PLACES_API_KEY else 'OFF'}")

    prospects = worker_dump("prospects", run_date)
    photos = worker_dump("photos", run_date)
    cards = worker_dump("cards", run_date)

    if SEND_MODE == "individual":
        if AGENCY not in VALID_AGENCIES:
            raise RuntimeError("AGENCY required (GR|VR|GRS|SLS) for mode=individual")
        if not INITIALS:
            raise RuntimeError("INITIALS required for mode=individual")

        send_individual_pack(run_date, AGENCY, INITIALS, prospects, photos, cards)
        print(json.dumps(STATS, ensure_ascii=False))
        return

    if SEND_MODE == "agency_manager":
        for ag in sorted(VALID_AGENCIES):
            send_agency_manager_pack(run_date, ag, prospects, photos, cards)
        print(json.dumps(STATS, ensure_ascii=False))
        return

    if SEND_MODE == "admin":
        send_admin_pack(run_date, prospects, photos, cards)
        print(json.dumps(STATS, ensure_ascii=False))
        return

    raise RuntimeError(f"Unknown SEND_MODE={SEND_MODE}")


if __name__ == "__main__":
    main()
