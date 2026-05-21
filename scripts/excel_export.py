"""
excel_export.py — Module export Excel prospection v8.4
Ajoute colonne NOM COMMERCIAL (enseigne visible, ex: Intermarché SUPER Varces)
"""

import io, os, datetime, shutil, subprocess
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

CRM_HEADERS = [
    "NOM", "NOM COMMERCIAL", "RUE", "CODE POSTAL", "VILLE",
    "Téléphone", "Téléphone (Portable)", "Mail générique",
    "SIRET", "NAF", "ACTIVITE", "EFFECTIF", "DATE CREATION", "CAPITAL",
    "SITE WEB", "PREFIXE", "INTERLOCUTEUR", "DIRIGEANT",
    "RESUME ENTRETIEN", "COMMANDE", "QUALIFICATION",
    "CARTE DE VISITE", "AGENCE", "INITIALS",
]

FIELD_MAP = {
    "NOM": "name", "NOM COMMERCIAL": "nom_commercial",
    "RUE": "address", "CODE POSTAL": "postal_code", "VILLE": "city",
    "Téléphone": "phone", "Téléphone (Portable)": "phone2",
    "Mail générique": "email", "SIRET": "siret", "NAF": "naf",
    "ACTIVITE": "activity_summary", "EFFECTIF": "effectif",
    "DATE CREATION": "date_creation", "CAPITAL": "capital",
    "SITE WEB": "website", "PREFIXE": "contact_civility",
    "INTERLOCUTEUR": "interlocuteur", "DIRIGEANT": "dirigeant",
    "RESUME ENTRETIEN": "resume", "COMMANDE": "commande",
    "QUALIFICATION": "qualification", "CARTE DE VISITE": "card_photo_url",
    "AGENCE": "agency", "INITIALS": "initials",
}

AGENCY_COLORS = {"GR": "FF8C00", "VR": "228B22", "GRS": "FFD700", "SLS": "1E90FF"}
AGENCY_TEXT_COLORS = {"GR": "FFFFFF", "VR": "FFFFFF", "GRS": "000000", "SLS": "FFFFFF"}
SCORE_FILLS = {"high": "FFFFFF", "medium": "FFFDE7", "low": "FFF0F0"}
ROW_HEIGHT_WITH_IMG = 60
IMG_MAX_W, IMG_MAX_H = 90, 55

COL_WIDTHS = {
    "NOM": 28, "NOM COMMERCIAL": 30, "RUE": 30, "CODE POSTAL": 9, "VILLE": 18,
    "Téléphone": 14, "Téléphone (Portable)": 14, "Mail générique": 30,
    "SIRET": 16, "NAF": 7, "ACTIVITE": 28, "EFFECTIF": 14,
    "DATE CREATION": 14, "CAPITAL": 12, "SITE WEB": 28, "PREFIXE": 8,
    "INTERLOCUTEUR": 22, "DIRIGEANT": 22, "RESUME ENTRETIEN": 40,
    "COMMANDE": 20, "QUALIFICATION": 12, "CARTE DE VISITE": 14,
    "AGENCE": 8, "INITIALS": 8,
}

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
_img_cache: Dict[str, bytes] = {}


def score_level(s): return "high" if s >= 60 else "medium" if s >= 30 else "low"
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def header_fill(): return PatternFill("solid", start_color="2C3E50", end_color="2C3E50")
def header_font(): return Font(name="Arial", bold=True, color="FFFFFF", size=10)
def cell_font(bold=False): return Font(name="Arial", size=9, bold=bold)
def centered(): return Alignment(horizontal="center", vertical="center")
def left_aligned(): return Alignment(horizontal="left", vertical="center")
def wrap_aligned(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def row_fill(score): return PatternFill("solid", start_color=SCORE_FILLS[score_level(score)], end_color=SCORE_FILLS[score_level(score)])
def agency_fill(ag): return PatternFill("solid", start_color=AGENCY_COLORS.get(ag.upper(),"EEEEEE"), end_color=AGENCY_COLORS.get(ag.upper(),"EEEEEE"))
def agency_font(ag): return Font(name="Arial", bold=True, size=9, color=AGENCY_TEXT_COLORS.get(ag.upper(),"000000"))


def fetch_image_bytes(url):
    if not url: return b""
    if url in _img_cache: return _img_cache[url]
    try:
        r = requests.get(url, headers={"User-Agent": "ProspectionBot/8.4"}, timeout=15)
        if r.status_code == 200:
            _img_cache[url] = r.content
            return r.content
    except: pass
    return b""


def fetch_image_from_file_id(file_id):
    if not file_id or not TELEGRAM_TOKEN: return b""
    try:
        r = requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getFile", json={"file_id": file_id}, timeout=15)
        j = r.json()
        if not j.get("ok"): return b""
        return fetch_image_bytes(f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{j['result']['file_path']}")
    except: return b""


def resize_image_bytes(img_bytes, max_w, max_h):
    if not img_bytes: return None
    try:
        img = PILImage.open(io.BytesIO(img_bytes)).convert("RGB")
        img.thumbnail((max_w, max_h), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()
    except: return None


def build_xl_image(img_bytes):
    resized = resize_image_bytes(img_bytes, IMG_MAX_W, IMG_MAX_H)
    if not resized: return None
    try:
        xl = XLImage(io.BytesIO(resized))
        xl.width, xl.height = IMG_MAX_W, IMG_MAX_H
        return xl
    except: return None


def get_card_image(row):
    for k in ["card_photo_url", "facade_photo_url"]:
        url = (row.get(k) or "").strip()
        if url:
            img = fetch_image_bytes(url)
            if img: return img
    for k in ["card_photo_file_id", "facade_photo_file_id"]:
        fid = (row.get(k) or "").strip()
        if fid:
            img = fetch_image_from_file_id(fid)
            if img: return img
    return b""


def build_sheet(ws, rows, with_images=True, sheet_title="DATA"):
    ws.title = sheet_title
    for ci, h in enumerate(CRM_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = header_font(), header_fill(), centered(), thin_border()
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CRM_HEADERS))}1"
    for ci, h in enumerate(CRM_HEADERS, 1):
        w = COL_WIDTHS.get(h, 15)
        if h == "CARTE DE VISITE" and with_images: w = max(w, int(IMG_MAX_W/6)+2)
        ws.column_dimensions[get_column_letter(ci)].width = w

    col_card = CRM_HEADERS.index("CARTE DE VISITE") + 1

    for ri, row in enumerate(rows, 2):
        score = int(row.get("_confidence", 0) or 0)
        ag = (row.get("agency") or "").upper().strip()
        rf = row_fill(score)
        if with_images: ws.row_dimensions[ri].height = ROW_HEIGHT_WITH_IMG

        for ci, h in enumerate(CRM_HEADERS, 1):
            val = str(row.get(FIELD_MAP.get(h, ""), "") or "").strip()
            c = ws.cell(row=ri, column=ci)
            c.border = thin_border()

            if h == "CARTE DE VISITE":
                if with_images: c.value, c.fill = "", rf
                else: c.value, c.font, c.alignment, c.fill = val, cell_font(), left_aligned(), rf
                continue
            if h == "QUALIFICATION":
                q = val.lower().strip()
                fills = {"chaud": "FFD0D0", "tiede": "FFF3CD", "tiède": "FFF3CD", "froid": "D0E8FF"}
                labels = {"chaud": "🔥 Chaud", "tiede": "📅 Tiède", "tiède": "📅 Tiède", "froid": "❄️ Froid"}
                c.value = labels.get(q, val)
                c.fill = PatternFill("solid", fgColor=fills.get(q, SCORE_FILLS[score_level(score)]))
                c.font = cell_font(bold=q in ("chaud", "froid"))
                c.alignment = centered()
                continue
            if h == "AGENCE":
                c.value, c.font, c.fill, c.alignment = ag, agency_font(ag), agency_fill(ag), centered()
                continue
            if h in ("RESUME ENTRETIEN", "COMMANDE"):
                c.value, c.font, c.alignment, c.fill = val, cell_font(), wrap_aligned(), rf
                continue
            if h == "NOM COMMERCIAL":
                c.value, c.fill = val, rf
                c.font = Font(name="Arial", size=9, italic=bool(val), color="555555" if val else "000000")
                c.alignment = left_aligned()
                continue
            c.value, c.font, c.alignment, c.fill = val, cell_font(), left_aligned(), rf

        if with_images:
            img_bytes = get_card_image(row)
            if img_bytes:
                xl = build_xl_image(img_bytes)
                if xl: ws.add_image(xl, f"{get_column_letter(col_card)}{ri}")
            else:
                url = str(row.get("card_photo_url") or "").strip()
                ws.cell(ri, col_card).value = url or "—"
                ws.cell(ri, col_card).font = cell_font()
                ws.cell(ri, col_card).alignment = left_aligned()

    tr = len(rows) + 2
    ws.cell(tr, 1, f"Total : {len(rows)} fiche(s)").font = Font(name="Arial", bold=True, size=9, italic=True)


def build_workbook_commercial(rows, with_images):
    wb = Workbook()
    build_sheet(wb.active, rows, with_images, "DATA")
    return wb


def build_workbook_agency(rows_by_initials, agency, with_images):
    wb = Workbook()
    all_rows = []
    for ini in sorted(rows_by_initials):
        all_rows.extend(rows_by_initials[ini])
    all_rows.sort(key=lambda r: int(r.get("_confidence", 0) or 0), reverse=True)
    build_sheet(wb.active, all_rows, with_images, agency)
    return wb


def build_workbook_admin(rows_by_agency, with_images):
    wb = Workbook()
    first = True
    for ag in ["GR", "VR", "GRS", "SLS"]:
        rows = rows_by_agency.get(ag, [])
        if not rows: continue
        rows_s = sorted(rows, key=lambda r: int(r.get("_confidence", 0) or 0), reverse=True)
        ws = wb.active if first else wb.create_sheet()
        first = False
        build_sheet(ws, rows_s, with_images, ag)
    return wb


def _try_libreoffice_convert(xlsx_path, out_dir, base_name):
    lo = shutil.which("libreoffice") or shutil.which("soffice")
    if not lo: return False
    try:
        r = subprocess.run([lo, "--headless", "--convert-to", "xls", "--outdir", out_dir, xlsx_path], capture_output=True, timeout=60)
        if r.returncode == 0:
            stem = os.path.splitext(os.path.basename(xlsx_path))[0]
            lo_out = os.path.join(out_dir, f"{stem}.xls")
            target = os.path.join(out_dir, f"{base_name}.xls")
            if os.path.exists(lo_out) and lo_out != target: os.rename(lo_out, target)
            return os.path.exists(target)
    except: pass
    return False


def save_pair(wb_cons, wb_crm, out_dir, base_name):
    import shutil as _shutil
    # Supprimer si c'est un fichier (pas un dossier)
    if os.path.exists(out_dir) and not os.path.isdir(out_dir):
        os.remove(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    path_xlsx = os.path.join(out_dir, f"{base_name}.xlsx")
    path_xls  = os.path.join(out_dir, f"{base_name}.xls")
    wb_cons.save(path_xlsx)
    tmp = os.path.join(out_dir, f"{base_name}_CRM_tmp.xlsx")
    wb_crm.save(tmp)
    if not _try_libreoffice_convert(tmp, out_dir, base_name): shutil.copy(tmp, path_xls)
    try: os.remove(tmp)
    except: pass
    return path_xlsx, path_xls


def export_commercial(date, agency, initials, rows, out_dir="out"):
    base = f"Prospection_{date}_{agency}_{initials}"
    return save_pair(build_workbook_commercial(rows, True), build_workbook_commercial(rows, False), out_dir, base)


def export_agency_manager(date, agency, rows, out_dir="out"):
    by_ini = {}
    for r in rows: by_ini.setdefault((r.get("initials") or "").upper(), []).append(r)
    base = f"Prospection_{date}_{agency}_MANAGER"
    return save_pair(build_workbook_agency(by_ini, agency, True), build_workbook_agency(by_ini, agency, False), out_dir, base)


def export_admin(date, rows_all, out_dir="out"):
    by_ag = {}
    for r in rows_all: by_ag.setdefault((r.get("agency") or "").upper(), []).append(r)
    base = f"Prospection_{date}_ADMIN"
    return save_pair(build_workbook_admin(by_ag, True), build_workbook_admin(by_ag, False), out_dir, base)


def get_xlsx_bytes(path):
    with open(path, "rb") as f: return f.read()


def build_email_attachments(path_xlsx, path_xls):
    return [(os.path.basename(p), get_xlsx_bytes(p)) for p in [path_xlsx, path_xls] if os.path.exists(p)]
